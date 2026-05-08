from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional
from datetime import datetime
import asyncio
from pydantic import BaseModel

from app.db import get_db
from app.dependencies import check_permission, get_current_user
from app.models.network import DiscoveryJob, DiscoveredDevice, DiscoveryLog
from app.models.hardware import PC, Monitor, DockingStation, Phone
from app.services.discovery_service import (
    run_discovery_job, resolve_hostnames_from_ips
)
from app.services.switch_telemetry import (
    get_switch_ports, get_arp_table, get_device_summary,
    get_device_hardware_software, get_device_power, get_device_fans,
    get_device_temperature, get_device_sfp, get_device_vlans,
    get_device_stp, get_device_stack
)

router = APIRouter(
    prefix="/discovery",
    tags=["discovery"],
    dependencies=[Depends(check_permission("network"))]
)

class JobCreate(BaseModel):
    name: str
    target_type: str
    target_value: str
    protocols: str

class DeviceUpdate(BaseModel):
    site: Optional[str] = None
    role: Optional[str] = None

@router.get("/stats")
async def get_discovery_stats(db: Session = Depends(get_db)):
    total = db.query(DiscoveredDevice).count()
    new_devs = db.query(DiscoveredDevice).filter(DiscoveredDevice.status == "NEW").count()
    unclassified = db.query(DiscoveredDevice).filter(DiscoveredDevice.device_type == "Unknown").count()
    failed_jobs = db.query(DiscoveryJob).filter(DiscoveryJob.status == "FAILED").count()
    
    last_job = db.query(DiscoveryJob).filter(DiscoveryJob.last_run != None).order_by(DiscoveryJob.last_run.desc()).first()
    
    return {
        "total_devices": total,
        "new_devices": new_devs,
        "unclassified": unclassified,
        "failed_jobs": failed_jobs,
        "last_scan": last_job.last_run.isoformat() if last_job and last_job.last_run else None
    }

@router.get("/devices")
def get_discovered_devices(db: Session = Depends(get_db)):
    devices = db.query(DiscoveredDevice).filter(DiscoveredDevice.status != 'IGNORED').order_by(DiscoveredDevice.last_seen.desc()).all()
    
    # Pre-fetch user mapping for all devices by MAC
    from sqlalchemy import text
    db_pcs = db.execute(text("""
        SELECT p.mac_address, u.full_name 
        FROM pcs p 
        LEFT JOIN users u ON p.current_user_id = u.id 
        WHERE p.mac_address IS NOT NULL AND u.id IS NOT NULL
    """)).fetchall()
    
    mac_to_user = {}
    for pc_mac, user_name in db_pcs:
        for mac in [m.strip().lower() for m in pc_mac.split(',')]:
            mac_to_user[mac] = user_name

    result = []
    for d in devices:
        user = None
        if d.mac_address:
            for mac in [m.strip().lower() for m in d.mac_address.split(',')]:
                if mac in mac_to_user:
                    user = mac_to_user[mac]
                    break
                    
        result.append({
            "id": d.id,
            "hostname": d.hostname,
            "ip_address": d.ip_address,
            "mac_address": d.mac_address,
            "device_type": d.device_type,
            "os_info": d.os_info,
            "vendor": d.vendor,
            "status": d.status,
            "last_seen": d.last_seen.isoformat() if d.last_seen else None,
            "matched_hw_id": d.matched_hw_id,
            "discovery_source": d.discovery_source,
            "snmp_status": d.snmp_status,
            "snmp_error": d.snmp_error,
            "serial_number": d.serial_number,
            "uptime": d.uptime,
            "assigned_user": user
        })
    return result

@router.get("/devices/export")
async def export_devices(ids: str = None, db: Session = Depends(get_db)):
    """Export all network devices to a professional Excel workbook with cached telemetry."""
    import io, json, asyncio
    from datetime import datetime
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from sqlalchemy import text
    from app.models.settings import SystemSetting
    from app.models.network import DeviceTelemetry
    
    devices_all = get_discovered_devices(db)
    
    # If explicit IDs were passed, filter to only those devices
    if ids:
        selected_ids = [int(i.strip()) for i in ids.split(',') if i.strip().isdigit()]
        devices_all = [d for d in devices_all if d.get("id") in selected_ids]
        
    # Filter out endpoints as requested
    devices = [d for d in devices_all if d.get("device_type") not in ("PC", "Server", "Printer")]
    
    # ── Collect DB lookups for endpoint resolution ──
    mac_to_ip, mac_to_hostname, mac_to_user = {}, {}, {}
    for fd in db.query(DiscoveredDevice).filter(DiscoveredDevice.mac_address.isnot(None)).all():
        if fd.mac_address:
            for mac in [m.strip().lower() for m in fd.mac_address.split(',')]:
                if fd.ip_address: mac_to_ip[mac] = fd.ip_address
                if fd.hostname and "HOST-" not in fd.hostname: mac_to_hostname[mac] = fd.hostname
    for pc_name, pc_ip, pc_mac, pc_user in db.execute(text("SELECT p.name,p.ip_address,p.mac_address,u.full_name FROM pcs p LEFT JOIN users u ON p.current_user_id=u.id WHERE p.mac_address IS NOT NULL")).fetchall():
        if pc_mac:
            for mac in [m.strip().lower() for m in pc_mac.split(',')]:
                if pc_ip and mac not in mac_to_ip: mac_to_ip[mac] = pc_ip
                if pc_name and mac not in mac_to_hostname: mac_to_hostname[mac] = pc_name
                if pc_user and mac not in mac_to_user: mac_to_user[mac] = pc_user
    for pr_model, pr_ip, pr_mac in db.execute(text("SELECT model,ip_address,mac_address FROM printers WHERE mac_address IS NOT NULL")).fetchall():
        if pr_mac:
            for mac in [m.strip().lower() for m in pr_mac.split(',')]:
                if pr_ip and mac not in mac_to_ip: mac_to_ip[mac] = pr_ip
                if pr_model and mac not in mac_to_hostname: mac_to_hostname[mac] = pr_model

    # ── Read cached telemetry (no live SNMP — instant) ──
    ports_list, power_list, vlan_list = [], [], []
    total_ports_up, total_ports_down, total_ports_admin = 0, 0, 0

    for d in devices:
        if d["device_type"] != "Switch" or d.get("snmp_status") != "CONNECTED":
            continue
        dev = db.query(DiscoveredDevice).filter(DiscoveredDevice.id == d["id"]).first()
        if not dev or not dev.telemetry: continue
        t = dev.telemetry
        sw_name = d["hostname"] or d["ip_address"]
        sw_ip = d["ip_address"]

        # Ports
        cached_ports = json.loads(t.ports_data_json) if t.ports_data_json else []
        cached_summary = json.loads(t.summary_data_json or "{}")
        cached_arp = cached_summary.get("arp", {})

        for p in cached_ports:
            admin = str(p.get("admin_status", ""))
            oper = str(p.get("oper_status", ""))
            admin_txt = "Up" if admin == "1" else ("Down" if admin == "2" else admin)
            if admin == "2": oper_txt = "Admin Down"; total_ports_admin += 1
            elif admin == "1" and oper == "1": oper_txt = "Connected"; total_ports_up += 1
            else: oper_txt = "Not Connected"; total_ports_down += 1

            speed_raw = int(p.get("speed", 0) or 0)
            if speed_raw >= 1e9: speed_txt = f"{int(speed_raw/1e9)} Gbps"
            elif speed_raw >= 1e6: speed_txt = f"{int(speed_raw/1e6)} Mbps"
            elif speed_raw > 0: speed_txt = f"{speed_raw} bps"
            else: speed_txt = "—"

            macs = p.get("macs_connected", [])
            ips, names, users = [], [], []
            if 0 < len(macs) <= 5:
                for m in macs:
                    if m in cached_arp: ips.append(cached_arp[m])
                    elif m in mac_to_ip: ips.append(mac_to_ip[m])
                    if m in mac_to_hostname: names.append(mac_to_hostname[m])
                    if m in mac_to_user: users.append(mac_to_user[m])
            end_ip = p.get("neighbor_ip") or ", ".join(list(dict.fromkeys([i.strip() for s in ips for i in s.split(',')])))
            end_name = p.get("neighbor_name") or ", ".join(list(dict.fromkeys(names)))
            end_user = ", ".join(list(dict.fromkeys(users)))
            raw_macs = ", ".join([m.upper() for m in macs[:5]])
            if len(macs) > 5: raw_macs += f" (+{len(macs)-5} more)"

            ports_list.append([sw_name, sw_ip, p.get("name",""), p.get("alias",""), admin_txt, oper_txt, speed_txt, p.get("vlan_id",""), raw_macs, end_ip, end_name, end_user, p.get("neighbor_caps","")])

        # Power / Fans / Environment
        pwr = cached_summary.get("power", {})
        fans_data = cached_summary.get("fans", {})
        env_data = cached_summary.get("environment", {})
        poe = pwr.get("poe") if pwr else None
        fan_str = " | ".join([f"{f['name']}: {f['state']}" for f in (fans_data.get("fans") or [])]) or "N/A"
        temp_str = " | ".join([f"{s['name']}: {s['value_celsius']}°C ({s['state']})" for s in (env_data.get("sensors") or [])]) or "N/A"
        psu_str = " | ".join([f"{s['name']}: {s['state']}" for s in (pwr.get("supplies") or [])]) or "N/A"
        power_list.append([sw_name, sw_ip, psu_str, poe.get("budget_watts","—") if poe else "—", poe.get("consumption_watts","—") if poe else "—", fan_str, temp_str])

        # VLANs
        for v in (cached_summary.get("vlans", {}).get("vlans") or []):
            vlan_list.append([sw_name, sw_ip, v.get("vlan_id",""), v.get("name",""), v.get("state",""), v.get("type","")])

    # ── Build workbook ──
    def generate_excel():
        wb = Workbook()
        # Style constants
        BRAND = "1F4E78"
        BRAND_FILL = PatternFill("solid", fgColor=BRAND)
        HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        TITLE_FONT = Font(bold=True, color=BRAND, size=16, name="Calibri")
        SUBTITLE_FONT = Font(color="666666", size=9, name="Calibri", italic=True)
        CELL_FONT = Font(size=10, name="Calibri")
        THIN_BORDER = Border(bottom=Side(style="thin", color="E0E0E0"))
        CENTER = Alignment(horizontal="center", vertical="center")
        LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
        GREEN_FILL = PatternFill("solid", fgColor="E8F5E9")
        RED_FILL = PatternFill("solid", fgColor="FFEBEE")
        GREY_FILL = PatternFill("solid", fgColor="F5F5F5")
        AMBER_FILL = PatternFill("solid", fgColor="FFF8E1")

        def style_sheet(ws, title, subtitle, headers, data, col_widths=None):
            ws.sheet_properties.tabColor = BRAND
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            ws['A1'] = title
            ws['A1'].font = TITLE_FONT
            ws['A1'].alignment = Alignment(vertical="center")
            ws.row_dimensions[1].height = 32
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
            ws['A2'] = subtitle
            ws['A2'].font = SUBTITLE_FONT
            ws.row_dimensions[2].height = 18
            # Header row (row 3)
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=3, column=ci, value=h)
                c.font = HEADER_FONT; c.fill = BRAND_FILL; c.alignment = CENTER
            ws.row_dimensions[3].height = 22
            ws.freeze_panes = "A4"
            # Data
            for ri, row in enumerate(data, 4):
                for ci, val in enumerate(row, 1):
                    c = ws.cell(row=ri, column=ci, value=val)
                    c.font = CELL_FONT; c.alignment = LEFT; c.border = THIN_BORDER
            # Column widths
            if col_widths:
                for ci, w in enumerate(col_widths, 1):
                    ws.column_dimensions[get_column_letter(ci)].width = w
            else:
                for ci in range(1, len(headers)+1):
                    mx = len(str(headers[ci-1]))
                    for row in data:
                        if ci-1 < len(row): mx = max(mx, len(str(row[ci-1] or "")))
                    ws.column_dimensions[get_column_letter(ci)].width = min(mx + 3, 45)
                    
            # Add auto-filter
            if data:
                last_col = get_column_letter(len(headers))
                last_row = 3 + len(data)
                ws.auto_filter.ref = f"A3:{last_col}{last_row}"
                
            return ws

        # ── Sheet 1: Dashboard ──
        ws0 = wb.active; ws0.title = "Dashboard"
        ws0.sheet_properties.tabColor = "0D47A1"
        ws0.merge_cells("A1:D1")
        ws0['A1'] = "KOSTAL ITSM — Network Discovery Report"
        ws0['A1'].font = Font(bold=True, color=BRAND, size=20, name="Calibri")
        ws0.row_dimensions[1].height = 40
        ws0['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws0['A2'].font = SUBTITLE_FONT
        # KPI cards
        kpis = [
            ("Total Devices", len(devices)),
            ("Switches (SNMP)", sum(1 for d in devices if d["device_type"]=="Switch" and d.get("snmp_status")=="CONNECTED")),
            ("Ports Connected", total_ports_up),
            ("Ports Down", total_ports_down),
            ("Ports Admin Down", total_ports_admin),
            ("Total Ports", total_ports_up + total_ports_down + total_ports_admin),
        ]
        for i, (label, value) in enumerate(kpis):
            r = 4 + i
            ws0.cell(row=r, column=1, value=label).font = Font(bold=True, size=11, name="Calibri", color="333333")
            vc = ws0.cell(row=r, column=2, value=value)
            vc.font = Font(bold=True, size=14, name="Calibri", color=BRAND)
            vc.alignment = CENTER
        for ci in [1,2,3,4]: ws0.column_dimensions[get_column_letter(ci)].width = 22

        # ── Sheet 2: Network Devices ──
        dev_headers = ["Hostname","IP Address","MAC Address","Type","Vendor","OS Info","Status","SNMP","Serial","Uptime","User","Last Seen"]
        dev_data = []
        for d in devices:
            dev_data.append([d["hostname"] or "Unknown", d["ip_address"] or "", d["mac_address"] or "", d["device_type"] or "", d["vendor"] or "", d["os_info"] or "", d["status"] or "", d["snmp_status"] or "", d["serial_number"] or "", d["uptime"] or "", d["assigned_user"] or "", d["last_seen"] or ""])
        ws1 = wb.create_sheet("Network Devices")
        style_sheet(ws1, "Network Devices Inventory", f"{len(dev_data)} devices discovered", dev_headers, dev_data, [22,16,20,12,14,30,12,14,18,22,18,20])
        # Color SNMP status
        for ri, row in enumerate(dev_data, 4):
            snmp_cell = ws1.cell(row=ri, column=8)
            if snmp_cell.value == "CONNECTED": snmp_cell.fill = GREEN_FILL
            elif snmp_cell.value == "FAILED": snmp_cell.fill = RED_FILL

        # ── Sheet 3: Switch Ports ──
        port_headers = ["Switch","Switch IP","Port","Description","Admin","Status","Speed","VLAN","MACs","Endpoint IP","Endpoint Name","User","CDP Caps"]
        ws2 = wb.create_sheet("Switch Ports")
        style_sheet(ws2, "Switch Port Inventory", f"{len(ports_list)} ports across {len(power_list)} switches", port_headers, ports_list, [20,16,22,24,10,16,12,8,22,16,22,18,14])
        for ri, row in enumerate(ports_list, 4):
            st_cell = ws2.cell(row=ri, column=6)
            if st_cell.value == "Connected": st_cell.fill = GREEN_FILL
            elif st_cell.value == "Admin Down": st_cell.fill = RED_FILL
            elif st_cell.value == "Not Connected": st_cell.fill = GREY_FILL

        # ── Sheet 4: VLANs ──
        vlan_headers = ["Switch","Switch IP","VLAN ID","VLAN Name","State","Type"]
        ws3 = wb.create_sheet("VLANs")
        style_sheet(ws3, "VLAN Configuration", f"{len(vlan_list)} VLANs across all switches", vlan_headers, vlan_list, [22,16,10,28,12,14])
        for ri, row in enumerate(vlan_list, 4):
            st_cell = ws3.cell(row=ri, column=5)
            if st_cell.value == "Active": st_cell.fill = GREEN_FILL
            elif st_cell.value == "Suspended": st_cell.fill = AMBER_FILL

        # ── Sheet 5: Environment & PoE ──
        env_headers = ["Switch","Switch IP","Power Supplies","PoE Budget (W)","PoE Used (W)","Fans","Temperature"]
        ws4 = wb.create_sheet("Environment & PoE")
        style_sheet(ws4, "Switch Power & Environment", f"{len(power_list)} switches", env_headers, power_list, [22,16,35,14,14,35,40])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    result_buf = await asyncio.to_thread(generate_excel)
    
    headers = {'Content-Disposition': 'attachment; filename="KOSTAL_Network_Discovery_Report.xlsx"'}
    return StreamingResponse(result_buf, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@router.get("/devices/{device_id}")
async def get_device_details(device_id: int, db: Session = Depends(get_db)):
    d = db.query(DiscoveredDevice).filter(DiscoveredDevice.id == device_id).first()
    if not d:
        raise HTTPException(status_code=404, detail="Device not found")
        
    return {
        "id": d.id,
        "hostname": d.hostname,
        "ip_address": d.ip_address,
        "mac_address": d.mac_address,
        "device_type": d.device_type,
        "os_info": d.os_info,
        "vendor": d.vendor,
        "snmp_status": d.snmp_status,
        "status": d.status,
        "last_seen": d.last_seen.isoformat() if d.last_seen else None,
        "last_polled_at": d.telemetry.last_polled_at.isoformat() if d.telemetry and d.telemetry.last_polled_at else None,
        "matched_hw_id": d.matched_hw_id,
        "discovery_source": d.discovery_source,
        "open_ports": d.open_ports,
        "serial_number": d.serial_number,
        "uptime": d.uptime,
        "site": d.site,
        "role": d.role
    }

@router.patch("/devices/{device_id}")
async def update_device(device_id: int, update_data: DeviceUpdate, db: Session = Depends(get_db)):
    d = db.query(DiscoveredDevice).filter(DiscoveredDevice.id == device_id).first()
    if not d:
        raise HTTPException(status_code=404, detail="Device not found")
        
    if update_data.site is not None:
        d.site = update_data.site
    if update_data.role is not None:
        d.role = update_data.role
        
    db.commit()
    return {"status": "success", "site": d.site, "role": d.role}

@router.get("/devices/{device_id}/ports")
async def get_device_ports(device_id: int, force: bool = False, db: Session = Depends(get_db)):
    d = db.query(DiscoveredDevice).filter(DiscoveredDevice.id == device_id).first()
    if not d:
        raise HTTPException(status_code=404, detail="Device not found")
    
    if d.snmp_status != "CONNECTED":
        raise HTTPException(status_code=400, detail="Device is not connected via SNMP.")
        
    ports = None
    arp_table = {}
    
    if not force and d.telemetry and d.telemetry.ports_data_json:
        import json
        ports = json.loads(d.telemetry.ports_data_json)
        summary = json.loads(d.telemetry.summary_data_json or "{}")
        arp_table = summary.get("arp", {})
    else:
        # Fetch global default community from SystemSetting
        from app.models.settings import SystemSetting
        global_community_setting = db.query(SystemSetting).filter(SystemSetting.setting_key == 'snmp_community').first()
        community = global_community_setting.setting_value if global_community_setting else "public"

        if d.discovery_source and "COMMUNITY:" in d.discovery_source.upper():
            parts = d.discovery_source.split(",")
            for p in parts:
                if p.upper().startswith("COMMUNITY:"):
                    community = p[len("COMMUNITY:"):]
                    break
                    
        ports = await get_switch_ports(d.ip_address, community)
        arp_table = await get_arp_table(d.ip_address, community)
    
    # Extract all unique MACs from the CAM table results
    all_macs = set()
    for p in ports:
        for m in p.get("macs_connected", []):
            all_macs.add(m)
            
    # Resolve IPs from ARP table
    snmp_mac_to_ip = {}
    snmp_ips_to_resolve = set()
    for mac in all_macs:
        if mac in arp_table:
            snmp_mac_to_ip[mac] = arp_table[mac]
            snmp_ips_to_resolve.add(arp_table[mac])
            
    # Reverse DNS
    snmp_ip_to_hostname = await resolve_hostnames_from_ips(list(snmp_ips_to_resolve))

    # Fetch database mappings to resolve IPs and hostnames from MACs
    mac_to_ip = {}
    mac_to_hostname = {}
    mac_to_user = {}
    mac_to_url = {}
    mac_to_user_url = {}
    
    # 1. Check Network Discovery table
    db_devices = db.query(DiscoveredDevice).filter(DiscoveredDevice.mac_address.isnot(None)).all()
    for fd in db_devices:
        if fd.mac_address:
            for mac in [m.strip().lower() for m in fd.mac_address.split(',')]:
                if fd.ip_address:
                    mac_to_ip[mac] = fd.ip_address
                if fd.hostname and "HOST-" not in fd.hostname:
                    mac_to_hostname[mac] = fd.hostname
                    
    # 2. Check Hardware Asset (PC) table using raw SQL to avoid mapper import errors
    from sqlalchemy import text
    db_pcs = db.execute(text("""
        SELECT p.id, p.name, p.ip_address, p.mac_address, u.full_name, u.id 
        FROM pcs p 
        LEFT JOIN users u ON p.current_user_id = u.id 
        WHERE p.mac_address IS NOT NULL
    """)).fetchall()
    for pc_id, pc_name, pc_ip, pc_mac, pc_user, u_id in db_pcs:
        if pc_mac:
            # PCs can sometimes have multiple MACs separated by comma
            for mac in [m.strip().lower() for m in pc_mac.split(',')]:
                if pc_ip and mac not in mac_to_ip:
                    mac_to_ip[mac] = pc_ip
                if pc_name and mac not in mac_to_hostname:
                    mac_to_hostname[mac] = pc_name
                if pc_user and mac not in mac_to_user:
                    mac_to_user[mac] = pc_user
                if mac not in mac_to_url:
                    mac_to_url[mac] = f"/hardware/pc/{pc_id}"
                if u_id and mac not in mac_to_user_url:
                    mac_to_user_url[mac] = f"/admin/users/{u_id}"

    # 3. Check Printers
    db_printers = db.execute(text("""
        SELECT id, model, ip_address, mac_address 
        FROM printers 
        WHERE mac_address IS NOT NULL
    """)).fetchall()
    for pr_id, pr_model, pr_ip, pr_mac in db_printers:
        if pr_mac:
            for mac in [m.strip().lower() for m in pr_mac.split(',')]:
                if pr_ip and mac not in mac_to_ip:
                    mac_to_ip[mac] = pr_ip
                if pr_model and mac not in mac_to_hostname:
                    mac_to_hostname[mac] = pr_model
                if mac not in mac_to_url:
                    mac_to_url[mac] = f"/hardware/printer/{pr_id}"
                
    # Inject IPs back into port data
    for p in ports:
        macs = p.get("macs_connected", [])
        p["data_source"] = "Unknown"
        
        # If CDP already populated neighbor_ip or neighbor_name, it's Live (SNMP)
        if p.get("neighbor_ip") or p.get("neighbor_name"):
            p["data_source"] = "Live (SNMP)"
            
        if len(macs) > 0 and len(macs) <= 5:
            # We skip massive trunks (like 200 MACs)
            ips = []
            names = []
            users = []
            urls = []
            user_urls = []
            sources = []
            
            for m in macs:
                # 1. Check SNMP/ARP first
                if m in snmp_mac_to_ip:
                    ip = snmp_mac_to_ip[m]
                    ips.append(ip)
                    if ip in snmp_ip_to_hostname:
                        names.append(snmp_ip_to_hostname[ip])
                    sources.append("Live (SNMP)")
                else:
                    # 2. DB Fallback for IP/Hostname
                    if m in mac_to_ip:
                        ips.append(mac_to_ip[m])
                        sources.append("ITSM Assets")
                    if m in mac_to_hostname:
                        names.append(mac_to_hostname[m])
                    if m in mac_to_url:
                        urls.append(mac_to_url[m])
                
                # DB is always the only source for users
                if m in mac_to_user:
                    users.append(mac_to_user[m])
                    sources.append("ITSM Assets")
                if m in mac_to_user_url:
                    user_urls.append(mac_to_user_url[m])
                    
            # If we didn't find CDP IP but we found end-user IPs, inject them
            if ips and not p.get("neighbor_ip"):
                p["neighbor_ip"] = ", ".join(list(dict.fromkeys([ip.strip() for ip_str in ips for ip in ip_str.split(',')])))
            if names and not p.get("neighbor_name"):
                p["neighbor_name"] = ", ".join(list(dict.fromkeys(names)))
            if urls and not p.get("neighbor_url"):
                p["neighbor_url"] = urls[0] # Take the first valid URL
            if users and not p.get("neighbor_user"):
                p["neighbor_user"] = ", ".join(list(dict.fromkeys(users)))
            if user_urls and not p.get("neighbor_user_url"):
                p["neighbor_user_url"] = user_urls[0]
                
            if "Live (SNMP)" in sources:
                p["data_source"] = "Live (SNMP)"
            elif "ITSM Assets" in sources:
                p["data_source"] = "ITSM Assets"
                
        # We can format the raw MACs into a clean string for the frontend
        if "macs_connected" in p and p["macs_connected"]:
            p["raw_macs"] = ", ".join([m.upper() for m in p["macs_connected"][:5]])
            if len(p["macs_connected"]) > 5:
                p["raw_macs"] += f" (+{len(p['macs_connected']) - 5} more)"
            del p["macs_connected"]
        else:
            p["raw_macs"] = ""

    return {"ports": ports}

@router.post("/devices/{device_id}/refresh")
async def refresh_device(device_id: int, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    d = db.query(DiscoveredDevice).filter(DiscoveredDevice.id == device_id).first()
    if not d:
        raise HTTPException(status_code=404, detail="Device not found")
        
    community = "public"
    if d.discovery_source:
        parts = d.discovery_source.split(",")
        for p in parts:
            if p.upper().startswith("COMMUNITY:"):
                community = p[len("COMMUNITY:"):]
                break
                
    from app.services.discovery_service import get_snmp_sysdescr
    from app.services.switch_telemetry import sync_poll_device_telemetry
    
    # Rapid SNMP check (skips slow ICMP and Port scanning for known devices)
    sys_descr, sys_name, snmp_status, snmp_error, uptime, serial_number = await get_snmp_sysdescr(d.ip_address, community)
    
    if snmp_status == "CONNECTED":
        d.snmp_status = "CONNECTED"
        if sys_name and "HOST-" not in sys_name:
            d.hostname = sys_name
        if sys_descr:
            d.device_type = "Network Switch"
            d.os_version = sys_descr[:100]
        if uptime:
            d.uptime = uptime
        if serial_number:
            d.serial_number = serial_number
    else:
        d.snmp_status = "FAILED"
        d.snmp_error = snmp_error
        
    d.last_seen = datetime.utcnow()
    db.commit()
        
    # Always poll telemetry after refresh to ensure cache is hot
    if d.snmp_status == "CONNECTED":
        # Run telemetry in background so UI returns instantly
        background_tasks.add_task(sync_poll_device_telemetry, d.id, d.ip_address, d.discovery_source)
        
    return {"message": "Refreshed"}

@router.post("/devices/{device_id}/sync")
async def sync_device_to_cmdb(device_id: int, db: Session = Depends(get_db)):
    d = db.query(DiscoveredDevice).filter(DiscoveredDevice.id == device_id).first()
    if not d:
        raise HTTPException(status_code=404, detail="Device not found")
        
    if d.status == "MATCHED":
        return {"message": "Already matched."}
        
    # Simple sync: Create a new PC if it's a Server or PC
    if d.device_type in ["PC", "Server", "Unknown"]:
        real_serial = d.serial_number or d.mac_address or f"IP-{d.ip_address}"
        
        # Check if exists by MAC or Serial
        existing = None
        if d.mac_address:
            existing = db.query(PC).filter(PC.mac_address == d.mac_address).first()
        if not existing:
            existing = db.query(PC).filter(PC.serial_number == real_serial).first()
            
        if existing:
            d.matched_hw_id = existing.id
            d.status = "MATCHED"
            db.commit()
            return {"message": "Matched to existing asset."}
                
        # create new
        new_pc = PC(
            name=d.hostname or f"Unknown-{d.ip_address}",
            serial_number=real_serial,
            vendor=d.vendor,
            ip_address=d.ip_address,
            mac_address=d.mac_address,
            windows_version=d.os_info,
            status="Available"
        )
        db.add(new_pc)
        db.commit()
        db.refresh(new_pc)
        
        d.matched_hw_id = new_pc.id
        d.status = "MATCHED"
        db.commit()
        return {"message": f"Created new asset ID {new_pc.id}"}
        
    return {"message": f"Sync not implemented for type {d.device_type}"}

@router.post("/devices/{device_id}/ignore")
async def ignore_device(device_id: int, db: Session = Depends(get_db)):
    d = db.query(DiscoveredDevice).filter(DiscoveredDevice.id == device_id).first()
    if not d:
        raise HTTPException(status_code=404)
    d.status = "IGNORED"
    db.commit()
    return {"message": "Device ignored"}

@router.get("/jobs")
async def list_jobs(db: Session = Depends(get_db)):
    jobs = db.query(DiscoveryJob).order_by(DiscoveryJob.created_at.desc()).all()
    return [{
        "id": j.id,
        "name": j.name,
        "target_type": j.target_type,
        "target_value": j.target_value,
        "protocols": j.protocols,
        "status": j.status,
        "last_run": j.last_run.isoformat() if j.last_run else None
    } for j in jobs]

@router.post("/jobs")
async def create_job(job: JobCreate, current_user = Depends(get_current_user), db: Session = Depends(get_db)):
    new_job = DiscoveryJob(
        name=job.name,
        target_type=job.target_type,
        target_value=job.target_value,
        protocols=job.protocols,
        created_by_id=current_user.id
    )
    db.add(new_job)
    db.commit()
    return {"id": new_job.id, "message": "Job created"}

@router.post("/jobs/{job_id}/run")
async def trigger_job(job_id: int, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    job = db.query(DiscoveryJob).filter(DiscoveryJob.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
        
    if job.status == "RUNNING":
        return {"message": "Job is already running"}
        
    job.status = "RUNNING"
    db.commit()
    
    background_tasks.add_task(run_discovery_job, job.id)
    return {"message": "Job started in background"}

@router.get("/locate_mac/{mac_address}")
async def locate_mac(mac_address: str, db: Session = Depends(get_db)):
    mac_address = mac_address.lower().strip()
    switches = db.query(DiscoveredDevice).filter(DiscoveredDevice.device_type == 'Switch').all()
    if not switches:
        return {"found": False, "message": "No switches configured in the network."}
    
    async def check_switch(switch):
        community = "public"
        if switch.discovery_source and "COMMUNITY:" in switch.discovery_source.upper():
            parts = switch.discovery_source.split(",")
            for p in parts:
                if p.upper().startswith("COMMUNITY:"):
                    community = p[len("COMMUNITY:"):]
                    break
        try:
            ports = await get_switch_ports(switch.ip_address, community)
            for p in ports:
                macs = [m.lower().strip() for m in p.get("macs_connected", [])]
                # Filter out trunk/uplink ports (if a port has many MACs, it's not the edge port)
                if mac_address in macs and len(macs) <= 5:
                    return {
                        "found": True, 
                        "switch_ip": switch.ip_address, 
                        "switch_name": switch.hostname, 
                        "port_name": p.get("name"), 
                        "port_status": p.get("status")
                    }
        except Exception as e:
            print(f"Exception in check_switch for {switch.ip_address}: {e}")
            pass
        return None

    results = await asyncio.gather(*[check_switch(sw) for sw in switches])
    for r in results:
        if r and r["found"]:
            return r
            
    return {"found": False, "message": "Device not currently active on known network switches."}


def _resolve_community(device, db):
    """Resolve SNMP community string from device or global settings."""
    from app.models.settings import SystemSetting
    global_community_setting = db.query(SystemSetting).filter(SystemSetting.setting_key == 'snmp_community').first()
    community = global_community_setting.setting_value if global_community_setting else "public"
    
    if device.discovery_source and "COMMUNITY:" in device.discovery_source.upper():
        parts = device.discovery_source.split(",")
        for p in parts:
            if p.upper().startswith("COMMUNITY:"):
                community = p[len("COMMUNITY:"):]
                break
    return community

@router.get("/devices/{device_id}/ports/{port_index}/traffic")
async def get_device_port_traffic(device_id: int, port_index: str, db: Session = Depends(get_db)):
    """Fetch live traffic for a specific port for charting."""
    d = db.query(DiscoveredDevice).filter(DiscoveredDevice.id == device_id).first()
    if not d:
        raise HTTPException(status_code=404, detail="Device not found")
        
    if d.snmp_status != "CONNECTED":
        raise HTTPException(status_code=400, detail="Device is not connected via SNMP.")
        
    community = _resolve_community(d, db)
    
    from app.services.switch_telemetry import get_port_traffic
    try:
        traffic_data = await get_port_traffic(d.ip_address, community, port_index)
        return traffic_data
    except Exception as e:
        logger.error(f"Error fetching port traffic for device {device_id} port {port_index}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/devices/{device_id}/ports/{port_index}/history")
async def get_device_port_history(device_id: int, port_index: str, db: Session = Depends(get_db)):
    """Fetch historical traffic and status for a specific port."""
    from app.models.network import PortTelemetryHistory
    from app.db import engine
    from datetime import datetime, timedelta
    import sqlalchemy
    
    try:
        # Auto-create table if user hasn't restarted
        PortTelemetryHistory.__table__.create(engine, checkfirst=True)
    except Exception:
        pass
    
    # Last 24 hours
    since = datetime.utcnow() - timedelta(hours=24)
    
    try:
        history_records = db.query(PortTelemetryHistory).filter(
            PortTelemetryHistory.device_id == device_id,
            PortTelemetryHistory.port_index == port_index,
            PortTelemetryHistory.timestamp >= since
        ).order_by(PortTelemetryHistory.timestamp.asc()).all()
    except sqlalchemy.exc.OperationalError:
        # Fallback if creation somehow failed
        history_records = []
    
    results = []
    last_in = None
    last_out = None
    last_time = None
    
    for r in history_records:
        mbps_in = 0
        mbps_out = 0
        
        if last_time and last_in is not None and last_out is not None:
            dt = (r.timestamp - last_time).total_seconds()
            if dt > 0:
                dIn = r.in_bytes - last_in
                if dIn < 0: dIn += (r.in_bytes > 0xFFFFFFFF and 0xFFFFFFFFFFFFFFFF or 0xFFFFFFFF)
                dOut = r.out_bytes - last_out
                if dOut < 0: dOut += (r.out_bytes > 0xFFFFFFFF and 0xFFFFFFFFFFFFFFFF or 0xFFFFFFFF)
                
                mbps_in = (dIn * 8) / (dt * 1000000)
                mbps_out = (dOut * 8) / (dt * 1000000)
                
        results.append({
            "timestamp": r.timestamp.isoformat() + "Z",
            "in_mbps": round(mbps_in, 2),
            "out_mbps": round(mbps_out, 2),
            "admin_status": r.admin_status,
            "oper_status": r.oper_status
        })
        
        last_in = r.in_bytes
        last_out = r.out_bytes
        last_time = r.timestamp
        
    return results

@router.get("/devices/{device_id}/detail/{section}")
async def get_device_section_data(device_id: int, section: str, force: bool = False, db: Session = Depends(get_db)):
    """
    Catalyst Center-style section data for Device 360.
    Section values: summary, hw-sw, power, fans, sfp, vlans, stp, stack, environment
    """
    d = db.query(DiscoveredDevice).filter(DiscoveredDevice.id == device_id).first()
    if not d:
        raise HTTPException(status_code=404, detail="Device not found")
    
    if d.snmp_status != "CONNECTED":
        raise HTTPException(status_code=400, detail="Device is not reachable via SNMP. Cannot fetch live data.")
    
    SECTION_MAP = {
        'summary': get_device_summary,
        'hw-sw': get_device_hardware_software,
        'power': get_device_power,
        'fans': get_device_fans,
        'sfp': get_device_sfp,
        'vlans': get_device_vlans,
        'stp': get_device_stp,
        'stack': get_device_stack,
        'environment': get_device_temperature,
    }
    
    handler = SECTION_MAP.get(section)
    if not handler:
        raise HTTPException(status_code=400, detail=f"Unknown section: {section}. Valid sections: {', '.join(SECTION_MAP.keys())}")

    # Check cache first
    if not force and d.telemetry and d.telemetry.summary_data_json:
        import json
        summary_data = json.loads(d.telemetry.summary_data_json)
        if section in summary_data:
            return {"section": section, "device_id": device_id, "data": summary_data[section], "cached": True}
            
    community = _resolve_community(d, db)
    ip = d.ip_address
    
    try:
        data = await handler(ip, community)
        return {"section": section, "device_id": device_id, "data": data, "cached": False}
    except Exception as e:
        logger.error(f"Error fetching section {section} for {ip}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/topology")
async def get_topology(db: Session = Depends(get_db)):
    """Fetch network topology for visualization."""
    from app.services.switch_telemetry import get_cdp_neighbors
    import asyncio
    
    devices = db.query(DiscoveredDevice).filter(DiscoveredDevice.status != 'IGNORED').all()
    if not devices:
        return {"nodes": [], "edges": []}
        
    nodes = []
    edges = []
    
    import json
    cdp_map = {}
    switches = [d for d in devices if d.device_type == 'Switch' and d.snmp_status == 'CONNECTED']
    
    for sw in switches:
        cdp_map[sw.ip_address] = []
        if sw.telemetry and sw.telemetry.summary_data_json:
            try:
                summary_data = json.loads(sw.telemetry.summary_data_json)
                if "cdp" in summary_data:
                    cdp_map[sw.ip_address] = summary_data["cdp"]
            except:
                pass
    
    ip_to_id = {d.ip_address: d.id for d in devices}
    added_edges = set()
    
    for d in devices:
        # Determine icon based on type
        group = "unknown"
        if d.device_type == "Switch":
            group = "switch"
        elif d.device_type == "Router":
            group = "router"
        elif d.device_type == "Firewall":
            group = "firewall"
        elif d.device_type == "Access Point":
            group = "ap"
        else:
            group = "endpoint"
            
        nodes.append({
            "id": d.id,
            "label": d.hostname or d.ip_address,
            "title": f"IP: {d.ip_address}<br>Vendor: {d.vendor or 'Unknown'}<br>Status: {d.snmp_status}",
            "group": group
        })
        
        # Add edges based on CDP
        if d.ip_address in cdp_map:
            for neighbor_ip in cdp_map[d.ip_address]:
                if neighbor_ip in ip_to_id:
                    neighbor_id = ip_to_id[neighbor_ip]
                    # Ensure unidirectional edge key to prevent duplicates
                    edge_key = tuple(sorted([d.id, neighbor_id]))
                    if edge_key not in added_edges:
                        edges.append({
                            "from": d.id,
                            "to": neighbor_id,
                            "color": {"color": "#4caf50"}
                        })
                        added_edges.add(edge_key)

    return {"nodes": nodes, "edges": edges}
