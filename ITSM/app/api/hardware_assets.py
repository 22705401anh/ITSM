import os
import uuid
import json
import shutil
import pandas as pd
from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, Query, Request
from fastapi.responses import JSONResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, and_, desc
from typing import List, Optional
from datetime import datetime

from app.db import get_db
from app.models.user import User
from app.models.hardware import PC, Monitor, DockingStation, Phone, PhoneNumber, Printer, AssetAssignment, StockAuditLog
from app.schemas.hardware import AssetAssignmentSchema, ReturnAssetSchema, AssetStatusUpdateSchema, DiscoveryPayloadSchema, AssetUpdateSchema

router = APIRouter(prefix="/hardware", tags=["Hardware Asset Tracking"])

# --- Upload directory setup ---
UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "uploads", "hardware")
os.makedirs(UPLOAD_DIR, exist_ok=True)

ALLOWED_IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp'}


def get_or_create_user(db: Session, full_name: str) -> User:
    user = db.query(User).filter(User.full_name == full_name).first()
    if not user:
        # Create a dummy user for now if missing, based on your logic
        user = User(
            username=full_name.lower().replace(" ", "_"), 
            email=f"{full_name.lower().replace(' ', '.')}@example.com",
            full_name=full_name,
            hashed_password="placeholder" # You will define proper auth
        )
        db.add(user)
        db.commit()
        db.refresh(user)
    return user


def log_audit(db: Session, action: str, asset_type: str, asset_id: int,
              serial_number: str = None, model: str = None,
              performed_by: str = "System", details: str = None):
    """Create an audit log entry."""
    entry = StockAuditLog(
        action=action,
        asset_type=asset_type,
        asset_id=asset_id,
        serial_number=serial_number,
        model=model,
        performed_by=performed_by,
        details=details,
        created_at=datetime.utcnow()
    )
    db.add(entry)
    db.flush()


def assign_hardware(db: Session, hardware, new_user: User, asset_type: str, notes: str = None):
    # 1. Close current assignment if exists
    current_assignment_query = db.query(AssetAssignment).filter(
        getattr(AssetAssignment, f"{asset_type}_id") == hardware.id,
        AssetAssignment.is_active == True
    )

    current_assignment = current_assignment_query.first()
    previous_user_id = None

    if current_assignment:
        if current_assignment.new_user_id == new_user.id:
            return  # Already assigned to this user

        # Prevent automated discovery ping-ponging (e.g. shared monitors / docking stations)
        if notes and "Auto-assigned" in notes:
            time_since = datetime.utcnow() - current_assignment.assigned_date
            # If it was assigned less than 24 hours ago, ignore rapid automated flip-flops
            if time_since.total_seconds() < 86400:
                return

        previous_user_id = current_assignment.new_user_id
        current_assignment.is_active = False
        current_assignment.returned_date = datetime.utcnow()

    # 2. Create new assignment history entry
    new_assignment = AssetAssignment(
        previous_user_id=previous_user_id,
        new_user_id=new_user.id,
        is_active=True,
        notes=notes
    )

    # Set the dynamic ID based on asset type
    setattr(new_assignment, f"{asset_type}_id", hardware.id)
    db.add(new_assignment)
    db.flush()

    # 3. Update current user on hardware table
    hardware.current_user_id = new_user.id
    hardware.status = "Assigned"
    db.add(hardware)
    db.flush()

    # 4. Log audit
    log_audit(db, "assigned", asset_type, hardware.id,
              serial_number=hardware.serial_number,
              model=hardware.model,
              details=f"Assigned to {new_user.full_name}" + (f" | Notes: {notes}" if notes else ""))


def return_hardware(db: Session, hardware, asset_type: str, notes: str = None):
    # 1. Close current assignment if exists
    current_assignment_query = db.query(AssetAssignment).filter(
        getattr(AssetAssignment, f"{asset_type}_id") == hardware.id,
        AssetAssignment.is_active == True
    )

    current_assignment = current_assignment_query.first()

    if current_assignment:
        current_assignment.is_active = False
        current_assignment.returned_date = datetime.utcnow()

    # 2. Update hardware status
    hardware.current_user_id = None
    hardware.status = "Available"
    db.add(hardware)
    db.flush()

    # 3. Log audit
    log_audit(db, "returned", asset_type, hardware.id,
              serial_number=hardware.serial_number,
              model=hardware.model,
              details=f"Returned to stock" + (f" | Notes: {notes}" if notes else ""))


@router.get("/search")
async def global_asset_search(query: str, db: Session = Depends(get_db)):
    """Searches across all asset tables simultaneously."""
    search_term = f"%{query}%"

    pcs = db.query(PC).outerjoin(User, PC.current_user_id == User.id).filter(or_(
        PC.serial_number.ilike(search_term), 
        PC.name.ilike(search_term), 
        PC.model.ilike(search_term),
        PC.ip_address.ilike(search_term),
        PC.mac_address.ilike(search_term),
        User.full_name.ilike(search_term),
        User.username.ilike(search_term)
    )).all()
    monitors = db.query(Monitor).outerjoin(User, Monitor.current_user_id == User.id).filter(or_(
        Monitor.serial_number.ilike(search_term), 
        Monitor.model.ilike(search_term),
        User.full_name.ilike(search_term),
        User.username.ilike(search_term)
    )).all()
    docks = db.query(DockingStation).outerjoin(User, DockingStation.current_user_id == User.id).filter(or_(
        DockingStation.serial_number.ilike(search_term), 
        DockingStation.model.ilike(search_term),
        User.full_name.ilike(search_term),
        User.username.ilike(search_term)
    )).all()
    phones = db.query(Phone).outerjoin(User, Phone.current_user_id == User.id).filter(or_(
        Phone.serial_number.ilike(search_term), 
        Phone.phone_number.ilike(search_term), 
        Phone.model.ilike(search_term),
        User.full_name.ilike(search_term),
        User.username.ilike(search_term)
    )).all()
    phone_numbers = db.query(PhoneNumber).outerjoin(User, PhoneNumber.current_user_id == User.id).filter(or_(
        PhoneNumber.serial_number.ilike(search_term), 
        PhoneNumber.phone_number.ilike(search_term), 
        PhoneNumber.model.ilike(search_term),
        User.full_name.ilike(search_term),
        User.username.ilike(search_term)
    )).all()

    return {
        "pcs": pcs,
        "monitors": monitors,
        "docking_stations": docks,
        "phones": phones,
        "phone_numbers": phone_numbers
    }


@router.get("/stock")
async def get_available_stock(db: Session = Depends(get_db)):
    """Fetch all available unassigned hardware with full details."""
    pcs = db.query(PC).filter(PC.status == "Available").all()
    monitors = db.query(Monitor).filter(Monitor.status == "Available").all()
    docks = db.query(DockingStation).filter(DockingStation.status == "Available").all()
    phones = db.query(Phone).filter(Phone.status == "Available").all()
    phone_numbers = db.query(PhoneNumber).filter(PhoneNumber.status == "Available").all()
    
    return {
        "pcs": [
            {
                "id": p.id,
                "asset_type": "pc",
                "name": p.name,
                "serial_number": p.serial_number,
                "model": p.model,
                "image_path": p.image_path,
                "notes": p.notes,
                "created_at": p.created_at.isoformat() if p.created_at else None,
                "status": p.status
            } for p in pcs
        ],
        "monitors": [
            {
                "id": m.id,
                "asset_type": "monitor",
                "serial_number": m.serial_number,
                "model": m.model,
                "image_path": m.image_path,
                "notes": m.notes,
                "created_at": m.created_at.isoformat() if m.created_at else None,
                "status": m.status
            } for m in monitors
        ],
        "docking_stations": [
            {
                "id": d.id,
                "asset_type": "docking_station",
                "serial_number": d.serial_number,
                "model": d.model,
                "image_path": d.image_path,
                "notes": d.notes,
                "created_at": d.created_at.isoformat() if d.created_at else None,
                "status": d.status
            } for d in docks
        ],
        "phones": [
            {
                "id": ph.id,
                "asset_type": "phone",
                "serial_number": ph.serial_number,
                "model": ph.model,
                "phone_number": ph.phone_number,
                "image_path": ph.image_path,
                "notes": ph.notes,
                "created_at": ph.created_at.isoformat() if ph.created_at else None,
                "status": ph.status
            } for ph in phones
        ],
        "phone_numbers": [
            {
                "id": pn.id,
                "asset_type": "phone_number",
                "serial_number": pn.serial_number,
                "model": pn.model,
                "phone_number": pn.phone_number,
                "image_path": pn.image_path,
                "notes": pn.notes,
                "created_at": pn.created_at.isoformat() if pn.created_at else None,
                "status": pn.status
            } for pn in phone_numbers
        ]
    }


@router.get("/stock/all")
async def get_all_stock(db: Session = Depends(get_db)):
    """Fetch ALL hardware (all statuses) for a unified view."""
    pcs = db.query(PC).all()
    monitors = db.query(Monitor).all()
    docks = db.query(DockingStation).all()
    phones = db.query(Phone).all()
    printers = db.query(Printer).all()
    phone_numbers = db.query(PhoneNumber).all()
    
    items = []
    for p in pcs:
        items.append({
            "id": p.id, "asset_type": "pc", "name": p.name,
            "serial_number": p.serial_number, "model": p.model,
            "vendor": getattr(p, "vendor", None),
            "ip_address": getattr(p, "ip_address", None),
            "mac_address": getattr(p, "mac_address", None),
            "last_seen_at": p.last_seen_at.isoformat() if getattr(p, "last_seen_at", None) else None,
            "image_path": p.image_path, "notes": p.notes, "status": p.status,
            "created_at": p.created_at.isoformat() if p.created_at else None,
            "assigned_user": p.current_user.full_name if p.current_user else None,
            "assigned_user_id": p.current_user_id
        })
    for m in monitors:
        items.append({
            "id": m.id, "asset_type": "monitor", "name": None,
            "serial_number": m.serial_number, "model": m.model,
            "vendor": getattr(m, "vendor", None),
            "last_seen_at": m.last_seen_at.isoformat() if getattr(m, "last_seen_at", None) else None,
            "image_path": m.image_path, "notes": m.notes, "status": m.status,
            "created_at": m.created_at.isoformat() if m.created_at else None,
            "assigned_user": m.current_user.full_name if m.current_user else None,
            "assigned_user_id": m.current_user_id
        })
    for d in docks:
        items.append({
            "id": d.id, "asset_type": "docking_station", "name": None,
            "serial_number": d.serial_number, "model": d.model,
            "image_path": d.image_path, "notes": d.notes, "status": d.status,
            "created_at": d.created_at.isoformat() if d.created_at else None,
            "assigned_user": d.current_user.full_name if d.current_user else None,
            "assigned_user_id": d.current_user_id
        })
    for ph in phones:
        items.append({
            "id": ph.id, "asset_type": "phone", "name": ph.phone_number,
            "serial_number": ph.serial_number, "model": ph.model,
            "image_path": ph.image_path, "notes": ph.notes, "status": ph.status,
            "created_at": ph.created_at.isoformat() if ph.created_at else None,
            "assigned_user": ph.current_user.full_name if ph.current_user else None,
            "assigned_user_id": ph.current_user_id
        })
    printers = db.query(Printer).all()
    for p in printers:
        items.append({
            "id": p.id, "asset_type": "printer", "name": p.name,
            "serial_number": p.serial_number, "model": p.model,
            "ip_address": p.ip_address,
            "mac_address": getattr(p, "mac_address", None),
            "last_seen_at": p.last_seen_at.isoformat() if getattr(p, "last_seen_at", None) else None,
            "image_path": p.image_path, "notes": p.notes, "status": p.status,
            "created_at": p.created_at.isoformat() if p.created_at else None,
            "assigned_user": p.current_user.full_name if p.current_user else None,
            "assigned_user_id": p.current_user_id
        })
    for pn in phone_numbers:
        items.append({
            "id": pn.id, "asset_type": "phone_number", "name": pn.phone_number,
            "serial_number": pn.serial_number, "model": pn.model,
            "image_path": pn.image_path, "notes": pn.notes, "status": pn.status,
            "created_at": pn.created_at.isoformat() if pn.created_at else None,
            "assigned_user": pn.current_user.full_name if pn.current_user else None,
            "assigned_user_id": pn.current_user_id
        })
    
    return {"items": items, "total": len(items)}

@router.get("/stock/export")
async def export_stock(db: Session = Depends(get_db)):
    """Export all hardware stock to a professional Excel Dashboard."""
    import io
    import pandas as pd
    import openpyxl
    from fastapi.responses import StreamingResponse
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    stock_data = await get_all_stock(db)
    items = stock_data["items"]
    
    export_list = []
    for item in items:
        export_list.append({
            "Asset Type": str(item.get("asset_type", "")).replace("_", " ").title(),
            "Name / Phone": item.get("name", ""),
            "Model": item.get("model", ""),
            "Serial Number": item.get("serial_number", ""),
            "Status": item.get("status", ""),
            "Assigned User": item.get("assigned_user", ""),
            "Vendor": item.get("vendor", ""),
            "IP Address": item.get("ip_address", ""),
            "MAC Address": item.get("mac_address", ""),
            "Last Seen": item.get("last_seen_at", ""),
            "Created At": item.get("created_at", ""),
            "Notes": item.get("notes", "")
        })
        
    df = pd.DataFrame(export_list)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 1. Write the raw data
        df.to_excel(writer, index=False, sheet_name='Data', startrow=1)
        workbook = writer.book
        data_sheet = writer.sheets['Data']
        
        # Format Data sheet
        data_sheet.title = "Asset Database"
        data_sheet.sheet_view.showGridLines = False
        
        data_sheet['A1'] = "Hardware Stock Database"
        data_sheet['A1'].font = Font(size=16, bold=True, color="1F4E78")
        
        for col_idx, column_cells in enumerate(data_sheet.columns, 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            data_sheet.column_dimensions[col_letter].width = min(max_length + 4, 45)
            
        data_sheet.freeze_panes = "A3"
        
        if len(df) > 0:
            tab = Table(displayName="HardwareStock", ref=f"A2:{data_sheet.cell(row=len(df)+2, column=len(df.columns)).coordinate}")
            style = TableStyleInfo(name="TableStyleLight8", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style
            data_sheet.add_table(tab)
            
            data_sheet.row_dimensions[2].height = 25
            for row in range(3, len(df)+3):
                data_sheet.row_dimensions[row].height = 22
                for col in range(1, len(df.columns) + 1):
                    data_sheet.cell(row=row, column=col).alignment = Alignment(vertical='center')
                    
        # 2. Create Executive Summary Dashboard Sheet
        summary_sheet = workbook.create_sheet('Executive Summary', 0)
        workbook.active = summary_sheet
        summary_sheet.sheet_view.showGridLines = False
        
        summary_sheet['B2'] = " KOSTAL ITSM"
        summary_sheet['B2'].font = Font(size=24, bold=True, color="FFFFFF")
        summary_sheet['B2'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        summary_sheet['B2'].alignment = Alignment(horizontal="left", vertical="center")
        
        summary_sheet['E2'] = "STOCK INVENTORY DASHBOARD "
        summary_sheet['E2'].font = Font(size=14, bold=True, color="FFFFFF")
        summary_sheet['E2'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        summary_sheet['E2'].alignment = Alignment(horizontal="right", vertical="center")
        
        summary_sheet.merge_cells('B2:D3')
        summary_sheet.merge_cells('E2:K3')
        
        from datetime import datetime
        summary_sheet['B4'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        summary_sheet['B4'].font = Font(italic=True, color="7F7F7F")
        
        kpi_titles = ["Total Hardware", "Available Stock", "Assigned in Use", "Maintenance / Broken"]
        
        total_assets = len(df)
        available = len(df[df['Status'].str.lower() == 'available']) if 'Status' in df else 0
        in_use = len(df[df['Status'].str.lower() == 'assigned']) if 'Status' in df else 0
        maintenance = len(df[df['Status'].str.lower().isin(['broken', 'maintenance'])]) if 'Status' in df else 0
        
        kpi_values = [total_assets, available, in_use, maintenance]
        cols = ['B', 'E', 'H', 'J']
        
        thin_border = Border(left=Side(style='thin', color="CCCCCC"), right=Side(style='thin', color="CCCCCC"), top=Side(style='thin', color="CCCCCC"), bottom=Side(style='thin', color="CCCCCC"))
        
        for i in range(4):
            col = cols[i]
            end_col = chr(ord(col) + 1) if i < 3 else 'K'
            
            summary_sheet[f'{col}6'] = kpi_titles[i]
            summary_sheet[f'{col}6'].font = Font(size=11, color="7F7F7F", bold=True)
            summary_sheet[f'{col}6'].alignment = Alignment(horizontal="center", vertical="center")
            summary_sheet[f'{col}6'].fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
            
            summary_sheet[f'{col}7'] = kpi_values[i]
            summary_sheet[f'{col}7'].font = Font(size=24, bold=True, color="2E75B6")
            summary_sheet[f'{col}7'].alignment = Alignment(horizontal="center", vertical="center")
            summary_sheet[f'{col}7'].fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
            
            summary_sheet.merge_cells(f'{col}6:{end_col}6')
            summary_sheet.merge_cells(f'{col}7:{end_col}8')
            
            for r_idx in range(6, 9):
                for c_idx in range(openpyxl.utils.column_index_from_string(col), openpyxl.utils.column_index_from_string(end_col)+1):
                    summary_sheet.cell(row=r_idx, column=c_idx).border = thin_border

        data_start_col = 100 
        status_counts = df['Status'].value_counts() if 'Status' in df else pd.Series()
        
        summary_sheet.cell(row=1, column=data_start_col).value = "Status"
        summary_sheet.cell(row=1, column=data_start_col+1).value = "Count"
        row_idx = 2
        for status, count in status_counts.items():
            summary_sheet.cell(row=row_idx, column=data_start_col).value = status if status else "Unknown"
            summary_sheet.cell(row=row_idx, column=data_start_col+1).value = count
            row_idx += 1
            
        if not status_counts.empty and row_idx > 2:
            pie = PieChart()
            pie.title = "Equipment Status"
            pie.style = 2  
            labels = Reference(summary_sheet, min_col=data_start_col, min_row=2, max_row=row_idx-1)
            data = Reference(summary_sheet, min_col=data_start_col+1, min_row=1, max_row=row_idx-1)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.dataLabels = DataLabelList()
            pie.dataLabels.showPercent = True
            pie.width = 14
            pie.height = 9
            summary_sheet.add_chart(pie, "B11")
            
        if 'Asset Type' in df:
            type_counts = df['Asset Type'].value_counts()
            summary_sheet.cell(row=1, column=data_start_col+3).value = "Asset Type"
            summary_sheet.cell(row=1, column=data_start_col+4).value = "Count"
            
            r2 = 2
            for cat, count in type_counts.items():
                summary_sheet.cell(row=r2, column=data_start_col+3).value = cat
                summary_sheet.cell(row=r2, column=data_start_col+4).value = count
                r2 += 1
                
            if not type_counts.empty and r2 > 2:
                bar = BarChart()
                bar.title = "Hardware Types"
                bar.style = 13
                labels2 = Reference(summary_sheet, min_col=data_start_col+3, min_row=2, max_row=r2-1)
                data2 = Reference(summary_sheet, min_col=data_start_col+4, min_row=1, max_row=r2-1)
                bar.add_data(data2, titles_from_data=True)
                bar.set_categories(labels2)
                bar.varyColors = True
                bar.legend = None
                bar.width = 14
                bar.height = 9
                summary_sheet.add_chart(bar, "G11")
            
        summary_sheet.column_dimensions['A'].width = 3
        summary_sheet.column_dimensions['B'].width = 15
        summary_sheet.column_dimensions['C'].width = 10
        summary_sheet.column_dimensions['D'].width = 15
        summary_sheet.column_dimensions['E'].width = 15
        summary_sheet.column_dimensions['F'].width = 10
        summary_sheet.column_dimensions['G'].width = 15
        summary_sheet.column_dimensions['H'].width = 15
        summary_sheet.column_dimensions['I'].width = 10
        summary_sheet.column_dimensions['J'].width = 15
        summary_sheet.column_dimensions['K'].width = 15
            
    output.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="hardware_stock_dashboard.xlsx"'
    }
    return StreamingResponse(
        output, 
        headers=headers, 
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@router.get("/stale")
def get_stale_hardware(db: Session = Depends(get_db)):
    """Fetch hardware that hasn't been seen on the network in over 14 days."""
    from datetime import timedelta
    cutoff = datetime.utcnow() - timedelta(days=14)
    
    stale = []
    
    # PCs
    pcs = db.query(PC).filter(or_(PC.last_seen_at < cutoff, and_(PC.last_seen_at == None, PC.created_at < cutoff))).all()
    for p in pcs:
        stale.append({"type": "pc", "id": p.id, "model": p.model, "serial_number": p.serial_number, "last_seen": p.last_seen_at.isoformat() if p.last_seen_at else None})
        
    # Monitors
    mons = db.query(Monitor).filter(or_(Monitor.last_seen_at < cutoff, and_(Monitor.last_seen_at == None, Monitor.created_at < cutoff))).all()
    for m in mons:
        stale.append({"type": "monitor", "id": m.id, "model": m.model, "serial_number": m.serial_number, "last_seen": m.last_seen_at.isoformat() if m.last_seen_at else None})
        
    # Docking Stations
    docks = db.query(DockingStation).filter(or_(DockingStation.last_seen_at < cutoff, and_(DockingStation.last_seen_at == None, DockingStation.created_at < cutoff))).all()
    for d in docks:
        stale.append({"type": "docking_station", "id": d.id, "model": d.model, "serial_number": d.serial_number, "last_seen": d.last_seen_at.isoformat() if d.last_seen_at else None})
        
    # Printers
    printers = db.query(Printer).filter(or_(Printer.last_seen_at < cutoff, and_(Printer.last_seen_at == None, Printer.created_at < cutoff))).all()
    for p in printers:
        stale.append({"type": "printer", "id": p.id, "model": p.model, "serial_number": p.serial_number, "last_seen": p.last_seen_at.isoformat() if p.last_seen_at else None})
        
    stale.sort(key=lambda x: x["last_seen"] if x["last_seen"] else "")

    return stale


@router.get("/detail/{asset_type}/{asset_id}")
async def get_hardware_detail(asset_type: str, asset_id: int, db: Session = Depends(get_db)):
    """Fetch details for a single hardware asset by type and ID."""
    type_map = {
        "pc": PC,
        "monitor": Monitor,
        "docking_station": DockingStation,
        "phone": Phone,
        "phone_number": PhoneNumber,
        "printer": Printer
    }
    
    if asset_type not in type_map:
        raise HTTPException(status_code=400, detail="Invalid asset type")
        
    model_class = type_map[asset_type]
    asset = db.query(model_class).filter(model_class.id == asset_id).first()
    
    if not asset:
        raise HTTPException(status_code=404, detail="Hardware asset not found")
        
    result = {
        "id": asset.id,
        "asset_type": asset_type,
        "serial_number": asset.serial_number,
        "model": asset.model,
        "status": asset.status,
        "notes": asset.notes,
        "created_at": asset.created_at.isoformat() if asset.created_at else None,
        "image_path": asset.image_path
    }
    
    if asset_type == "pc":
        result.update({
            "name": asset.name,
            "vendor": getattr(asset, "vendor", None),
            "ip_address": getattr(asset, "ip_address", None),
            "mac_address": getattr(asset, "mac_address", None),
            "last_seen_at": asset.last_seen_at.isoformat() if getattr(asset, "last_seen_at", None) else None,
            "windows_version": getattr(asset, "windows_version", None),
            "intune_status": getattr(asset, "intune_status", None),
            "antivirus_status": getattr(asset, "antivirus_status", None),
            "ram": getattr(asset, "ram", None),
            "storage": getattr(asset, "storage", None),
            "print_volume_30d": getattr(asset, "print_volume_30d", 0),
        })
        
        # Get connected monitors
        connected_mons = db.query(Monitor).filter(Monitor.connected_pc_id == asset.id).all()
        result["connected_monitors"] = [
            {"id": m.id, "model": m.model, "serial_number": m.serial_number, "vendor": m.vendor} for m in connected_mons
        ]
        
        # Get software history
        from app.models.hardware import InstalledSoftware
        software_records = db.query(InstalledSoftware).filter(InstalledSoftware.pc_id == asset.id).all()
        result["software"] = [
            {"name": s.name, "version": s.version, "installed_date": s.installed_date.isoformat() if s.installed_date else None, 
             "is_active": s.is_active, "removed_date": s.removed_date.isoformat() if s.removed_date else None} 
            for s in software_records
        ]
        
        # Get printers
        from app.models.hardware import InstalledPrinter
        printers = db.query(InstalledPrinter).filter(InstalledPrinter.pc_id == asset.id).all()
        result["printers"] = [
            {"name": p.name, "driver_name": p.driver_name, "port_name": p.port_name, 
             "is_network": p.is_network, "is_default": p.is_default}
            for p in printers
        ]
    elif asset_type == "printer":
        last_seen = getattr(asset, "last_seen_at", None)
        result.update({
            "name": getattr(asset, "name", None),
            "ip_address": getattr(asset, "ip_address", None),
            "mac_address": getattr(asset, "mac_address", None),
            "last_seen_at": last_seen.isoformat() if last_seen else None,
        })
    elif asset_type == "phone" or asset_type == "phone_number":
        result.update({
            "phone_number": asset.phone_number
        })
    elif asset_type == "monitor" or asset_type == "docking_station":
        last_seen = getattr(asset, "last_seen_at", None)
        result.update({
            "vendor": getattr(asset, "vendor", None),
            "last_seen_at": last_seen.isoformat() if last_seen else None,
        })
        
        if asset_type == "monitor" and getattr(asset, "connected_pc_id", None):
            pc = db.query(PC).filter(PC.id == asset.connected_pc_id).first()
            if pc:
                result["connected_pc"] = {
                    "id": pc.id,
                    "model": pc.model,
                    "serial_number": pc.serial_number,
                    "vendor": pc.vendor
                }
        
    # Get current assignee if any
    if asset.current_user_id:
        user = db.query(User).filter(User.id == asset.current_user_id).first()
        if user:
            result["current_assignee"] = {
                "id": user.id,
                "name": user.full_name or user.username,
                "email": user.email
            }
            
    return result

from pydantic import BaseModel
class StockCreateSchema(BaseModel):
    asset_type: str
    model: str
    serial_number: str
    name: Optional[str] = None
    phone_number: Optional[str] = None
    notes: Optional[str] = None

@router.post("/stock")
async def add_stock(item: StockCreateSchema, db: Session = Depends(get_db)):
    """Add a new device directly to available stock."""
    type_map = {
        "pc": PC,
        "monitor": Monitor,
        "docking_station": DockingStation,
        "phone": Phone,
        "phone_number": PhoneNumber,
        "printer": Printer
    }
    
    if item.asset_type not in type_map:
        raise HTTPException(status_code=400, detail="Invalid asset type")
        
    model_class = type_map[item.asset_type]
    
    # Check duplicate serial
    existing = db.query(model_class).filter(model_class.serial_number == item.serial_number).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Serial number {item.serial_number} already exists")

    new_device = model_class(
        serial_number=item.serial_number,
        model=item.model,
        status="Available",
        notes=item.notes,
        created_at=datetime.utcnow()
    )
    
    if item.asset_type == "pc":
        new_device.name = item.name or ""
    elif item.asset_type == "phone" or item.asset_type == "phone_number":
        new_device.phone_number = item.phone_number or ""
        
    db.add(new_device)
    db.flush()

    # Log audit
    log_audit(db, "created", item.asset_type, new_device.id,
              serial_number=item.serial_number,
              model=item.model,
              details=f"Manually added to stock. Model: {item.model}, S/N: {item.serial_number}")

    db.commit()
    return {"message": "Item added to stock successfully", "id": new_device.id}


@router.get("/history/{asset_type}/{asset_id}")
async def get_asset_history(asset_type: str, asset_id: int, db: Session = Depends(get_db)):
    """Fetch the assignment timeline of any specific asset type."""
    valid_types = ["pc", "monitor", "docking_station", "phone", "phone_number", "printer"]
    if asset_type not in valid_types:
        raise HTTPException(status_code=400, detail="Invalid asset type")

    from sqlalchemy.orm import joinedload
    
    filter_attr = getattr(AssetAssignment, f"{asset_type}_id")
    history = (db.query(AssetAssignment)
                 .options(joinedload(AssetAssignment.new_user), joinedload(AssetAssignment.previous_user))
                 .filter(filter_attr == asset_id)
                 .order_by(AssetAssignment.assigned_date.desc())
                 .all())

    results = []
    for h in history:
        results.append({
            "id": h.id,
            "assigned_date": h.assigned_date.isoformat() if h.assigned_date else None,
            "returned_date": h.returned_date.isoformat() if h.returned_date else None,
            "is_active": h.is_active,
            "notes": h.notes,
            "new_user_name": h.new_user.full_name if h.new_user else "Unknown",
            "previous_user_name": h.previous_user.full_name if h.previous_user else "Inventory Stock"
        })

    return results


@router.post("/assign")
async def assign_specific_hardware(payload: AssetAssignmentSchema, db: Session = Depends(get_db)):
    """Creates a new timeline record and updates the device's current_user_id"""
    user = db.query(User).filter(User.id == payload.user_id).first()
    if not user:
        raise HTTPException(404, "User not found")

    hardware = None
    if payload.asset_type == "pc":
        hardware = db.query(PC).filter(PC.id == payload.asset_id).first()
    elif payload.asset_type == "monitor":
        hardware = db.query(Monitor).filter(Monitor.id == payload.asset_id).first()
    elif payload.asset_type == "docking_station":
        hardware = db.query(DockingStation).filter(DockingStation.id == payload.asset_id).first()
    elif payload.asset_type == "phone":
        hardware = db.query(Phone).filter(Phone.id == payload.asset_id).first()
    elif payload.asset_type == "phone_number":
        hardware = db.query(PhoneNumber).filter(PhoneNumber.id == payload.asset_id).first()
    elif payload.asset_type == "printer":
        hardware = db.query(Printer).filter(Printer.id == payload.asset_id).first()
    elif payload.asset_type == "printer":
        hardware = db.query(Printer).filter(Printer.id == payload.asset_id).first()

    if not hardware:
        raise HTTPException(404, "Hardware asset not found")

    assign_hardware(db, hardware, user, payload.asset_type, payload.notes)

    db.commit()
    return {"message": "Hardware assigned successfully and history logged."}


@router.post("/return")
async def return_specific_hardware(payload: ReturnAssetSchema, db: Session = Depends(get_db)):
    """Closes active assignment and returns device to Available stock."""
    hardware = None
    if payload.asset_type == "pc":
        hardware = db.query(PC).filter(PC.id == payload.asset_id).first()
    elif payload.asset_type == "monitor":
        hardware = db.query(Monitor).filter(Monitor.id == payload.asset_id).first()
    elif payload.asset_type == "docking_station":
        hardware = db.query(DockingStation).filter(DockingStation.id == payload.asset_id).first()
    elif payload.asset_type == "phone":
        hardware = db.query(Phone).filter(Phone.id == payload.asset_id).first()
    elif payload.asset_type == "phone_number":
        hardware = db.query(PhoneNumber).filter(PhoneNumber.id == payload.asset_id).first()
    elif payload.asset_type == "printer":
        hardware = db.query(Printer).filter(Printer.id == payload.asset_id).first()
    elif payload.asset_type == "printer":
        hardware = db.query(Printer).filter(Printer.id == payload.asset_id).first()

    if not hardware:
        raise HTTPException(404, "Hardware asset not found")

    return_hardware(db, hardware, payload.asset_type, payload.notes)
    for printer in payload.network_printers:
        p_db = db.query(Printer).filter(Printer.serial_number == printer.serial_number).first()
        is_new = p_db is None
        if is_new:
            p_db = Printer(
                serial_number=printer.serial_number,
                model=printer.model,
                ip_address=printer.ip_address,
                mac_address=printer.mac_address,
                status="Available",
                created_at=datetime.utcnow(),
                last_seen_at=datetime.utcnow()
            )
            db.add(p_db)
            db.flush()
            created += 1
            log_audit(db, "created", "printer", p_db.id, serial_number=p_db.serial_number, model=p_db.model, details="Discovered on network")
        else:
            p_db.last_seen_at = datetime.utcnow()
            if printer.model: p_db.model = printer.model
            if printer.ip_address: p_db.ip_address = printer.ip_address
            if printer.mac_address: p_db.mac_address = printer.mac_address
            updated += 1
            log_audit(db, "updated", "printer", p_db.id, serial_number=p_db.serial_number, model=p_db.model, details="Updated via network discovery")

    db.commit()
    return {"message": "Hardware returned successfully."}


@router.put("/stock/{asset_type}/{asset_id}/status")
async def update_asset_status(asset_type: str, asset_id: int, payload: AssetStatusUpdateSchema, db: Session = Depends(get_db)):
    """Update the status of an asset manually."""
    valid_types = ["pc", "monitor", "docking_station", "phone", "phone_number", "printer"]
    if asset_type not in valid_types:
        raise HTTPException(status_code=400, detail="Invalid asset type")

    hardware = None
    if asset_type == "pc":
        hardware = db.query(PC).filter(PC.id == asset_id).first()
    elif asset_type == "monitor":
        hardware = db.query(Monitor).filter(Monitor.id == asset_id).first()
    elif asset_type == "docking_station":
        hardware = db.query(DockingStation).filter(DockingStation.id == asset_id).first()
    elif asset_type == "phone" or asset_type == "phone_number":
        hardware = db.query(Phone).filter(Phone.id == asset_id).first()
    elif asset_type == "printer":
        hardware = db.query(Printer).filter(Printer.id == asset_id).first()

    if not hardware:
        raise HTTPException(404, "Hardware asset not found")

    old_status = hardware.status
    hardware.status = payload.status
    db.flush()

    log_audit(db, "updated", asset_type, hardware.id,
              serial_number=hardware.serial_number,
              model=hardware.model,
              details=f"Status changed from {old_status} to {payload.status}")

    for printer in payload.network_printers:
        p_db = db.query(Printer).filter(Printer.serial_number == printer.serial_number).first()
        is_new = p_db is None
        if is_new:
            p_db = Printer(
                serial_number=printer.serial_number,
                model=printer.model,
                ip_address=printer.ip_address,
                mac_address=printer.mac_address,
                status="Available",
                created_at=datetime.utcnow(),
                last_seen_at=datetime.utcnow()
            )
            db.add(p_db)
            db.flush()
            created += 1
            log_audit(db, "created", "printer", p_db.id, serial_number=p_db.serial_number, model=p_db.model, details="Discovered on network")
        else:
            p_db.last_seen_at = datetime.utcnow()
            if printer.model: p_db.model = printer.model
            if printer.ip_address: p_db.ip_address = printer.ip_address
            if printer.mac_address: p_db.mac_address = printer.mac_address
            updated += 1
            log_audit(db, "updated", "printer", p_db.id, serial_number=p_db.serial_number, model=p_db.model, details="Updated via network discovery")

    db.commit()
    return {"message": f"Asset status updated to {payload.status}"}


@router.post("/import")
async def import_assets_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Validates and processes an Excel upload to upsert assets and generate history timeline."""

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Only Excel files are supported.")

    try:
        df = pd.read_excel(file.file)
    except Exception as e:
        raise HTTPException(400, f"Error reading Excel file: {str(e)}")

    # 2. Validate Required Columns
    expected_columns = [
        "user_name", "pc_name", "pc_serial_number", "pc_model", 
        "assigned_status", "monitor_model", "monitor_serial_number", 
        "docking_station_model", "docking_station_serial_number", 
        "phone_model", "phone_serial_number", "phone_number", "notes"
    ]

    # Make sure all columns exist (even if empty in data) to avoid pandas KeyErrors
    for col in expected_columns:
        if col not in df.columns:
            df[col] = None

    processed_count = 0
    created_count = 0
    skipped_count = 0

    for index, row in df.iterrows():
        # Clean data (convert NaN to None)
        user_name = None if pd.isna(row['user_name']) else str(row['user_name']).strip()
        status = "Available" if pd.isna(row['assigned_status']) else str(row['assigned_status']).strip()
        notes = None if pd.isna(row['notes']) else str(row['notes']).strip()

        user = None
        if user_name:
            user = get_or_create_user(db, user_name)

        # Process PC
        if pd.notna(row['pc_serial_number']):
            sn = str(row['pc_serial_number']).strip()
            pc = db.query(PC).filter(PC.serial_number == sn).first()
            is_new = pc is None
            if not pc:
                pc = PC(
                    serial_number=sn, 
                    name=str(row.get('pc_name', '')) if pd.notna(row.get('pc_name')) else None,
                    model=str(row.get('pc_model', '')) if pd.notna(row.get('pc_model')) else None,
                    status=status,
                    created_at=datetime.utcnow()
                )
                db.add(pc)
                db.flush()
                created_count += 1
            
            log_audit(db, "imported" if is_new else "updated", "pc", pc.id,
                      serial_number=sn,
                      model=pc.model,
                      details=f"{'Created' if is_new else 'Updated'} via Excel import (row {index + 1})")

            if user:
                assign_hardware(db, pc, user, "pc", notes)

        # Process Monitor
        if pd.notna(row['monitor_serial_number']):
            sn = str(row['monitor_serial_number']).strip()
            monitor = db.query(Monitor).filter(Monitor.serial_number == sn).first()
            is_new = monitor is None
            if not monitor:
                monitor = Monitor(
                    serial_number=sn, 
                    model=str(row.get('monitor_model', '')) if pd.notna(row.get('monitor_model')) else None,
                    status=status,
                    created_at=datetime.utcnow()
                )
                db.add(monitor)
                db.flush()
                created_count += 1

            log_audit(db, "imported" if is_new else "updated", "monitor", monitor.id,
                      serial_number=sn,
                      model=monitor.model,
                      details=f"{'Created' if is_new else 'Updated'} via Excel import (row {index + 1})")

            if user:
                assign_hardware(db, monitor, user, "monitor", notes)

        # Process Docking Station
        if pd.notna(row['docking_station_serial_number']):
            sn = str(row['docking_station_serial_number']).strip()
            dock = db.query(DockingStation).filter(DockingStation.serial_number == sn).first()
            is_new = dock is None
            if not dock:
                dock = DockingStation(
                    serial_number=sn, 
                    model=str(row.get('docking_station_model', '')) if pd.notna(row.get('docking_station_model')) else None,
                    status=status,
                    created_at=datetime.utcnow()
                )
                db.add(dock)
                db.flush()
                created_count += 1

            log_audit(db, "imported" if is_new else "updated", "docking_station", dock.id,
                      serial_number=sn,
                      model=dock.model,
                      details=f"{'Created' if is_new else 'Updated'} via Excel import (row {index + 1})")

            if user:
                assign_hardware(db, dock, user, "docking_station", notes)

        # Process Phone
        if pd.notna(row['phone_serial_number']):
            sn = str(row['phone_serial_number']).strip()
            phone = db.query(Phone).filter(Phone.serial_number == sn).first()
            is_new = phone is None
            if not phone:
                phone = Phone(
                    serial_number=sn, 
                    model=str(row.get('phone_model', '')) if pd.notna(row.get('phone_model')) else None,
                    phone_number=str(row.get('phone_number', '')) if pd.notna(row.get('phone_number')) else None,
                    status=status,
                    created_at=datetime.utcnow()
                )
                db.add(phone)
                db.flush()
                created_count += 1

            log_audit(db, "imported" if is_new else "updated", "phone", phone.id,
                      serial_number=sn,
                      model=phone.model,
                      details=f"{'Created' if is_new else 'Updated'} via Excel import (row {index + 1})")

            if user:
                assign_hardware(db, phone, user, "phone", notes)

        processed_count += 1

    for printer in payload.network_printers:
        p_db = db.query(Printer).filter(Printer.serial_number == printer.serial_number).first()
        is_new = p_db is None
        if is_new:
            p_db = Printer(
                serial_number=printer.serial_number,
                model=printer.model,
                ip_address=printer.ip_address,
                mac_address=printer.mac_address,
                status="Available",
                created_at=datetime.utcnow(),
                last_seen_at=datetime.utcnow()
            )
            db.add(p_db)
            db.flush()
            created += 1
            log_audit(db, "created", "printer", p_db.id, serial_number=p_db.serial_number, model=p_db.model, details="Discovered on network")
        else:
            p_db.last_seen_at = datetime.utcnow()
            if printer.model: p_db.model = printer.model
            if printer.ip_address: p_db.ip_address = printer.ip_address
            if printer.mac_address: p_db.mac_address = printer.mac_address
            updated += 1
            log_audit(db, "updated", "printer", p_db.id, serial_number=p_db.serial_number, model=p_db.model, details="Updated via network discovery")

    db.commit()
    return {
        "message": f"Successfully processed {processed_count} rows from Excel.",
        "processed": processed_count,
        "created": created_count,
    }


# --- Image Upload Endpoints ---

@router.post("/stock/{asset_type}/{asset_id}/image")
async def upload_asset_image(asset_type: str, asset_id: int,
                             file: UploadFile = File(...),
                             db: Session = Depends(get_db)):
    """Upload an image for a specific hardware asset."""
    type_map = {"pc": PC, "monitor": Monitor, "docking_station": DockingStation, "phone": Phone,
        "phone_number": PhoneNumber,
        "printer": Printer}
    
    if asset_type not in type_map:
        raise HTTPException(400, "Invalid asset type")
    
    model_class = type_map[asset_type]
    asset = db.query(model_class).filter(model_class.id == asset_id).first()
    if not asset:
        raise HTTPException(404, "Asset not found")
    
    # Validate file extension
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_IMAGE_EXTENSIONS:
        raise HTTPException(400, f"Invalid image format. Allowed: {', '.join(ALLOWED_IMAGE_EXTENSIONS)}")
    
    # Generate unique filename
    filename = f"{asset_type}_{asset_id}_{uuid.uuid4().hex[:8]}{ext}"
    filepath = os.path.join(UPLOAD_DIR, filename)
    
    # Delete old image if exists
    if asset.image_path:
        old_path = os.path.join(UPLOAD_DIR, os.path.basename(asset.image_path))
        if os.path.exists(old_path):
            os.remove(old_path)
    
    # Save new image
    with open(filepath, "wb") as f:
        shutil.copyfileobj(file.file, f)
    
    # Update asset
    asset.image_path = f"/uploads/hardware/{filename}"
    
    log_audit(db, "image_uploaded", asset_type, asset_id,
              serial_number=asset.serial_number,
              model=asset.model,
              details=f"Image uploaded: {file.filename}")
    
    for printer in payload.network_printers:
        p_db = db.query(Printer).filter(Printer.serial_number == printer.serial_number).first()
        is_new = p_db is None
        if is_new:
            p_db = Printer(
                serial_number=printer.serial_number,
                model=printer.model,
                ip_address=printer.ip_address,
                mac_address=printer.mac_address,
                status="Available",
                created_at=datetime.utcnow(),
                last_seen_at=datetime.utcnow()
            )
            db.add(p_db)
            db.flush()
            created += 1
            log_audit(db, "created", "printer", p_db.id, serial_number=p_db.serial_number, model=p_db.model, details="Discovered on network")
        else:
            p_db.last_seen_at = datetime.utcnow()
            if printer.model: p_db.model = printer.model
            if printer.ip_address: p_db.ip_address = printer.ip_address
            if printer.mac_address: p_db.mac_address = printer.mac_address
            updated += 1
            log_audit(db, "updated", "printer", p_db.id, serial_number=p_db.serial_number, model=p_db.model, details="Updated via network discovery")

    db.commit()
    return {"message": "Image uploaded successfully", "image_path": asset.image_path}


@router.delete("/stock/{asset_type}/{asset_id}/image")
async def delete_asset_image(asset_type: str, asset_id: int,
                             db: Session = Depends(get_db)):
    """Remove image from a hardware asset."""
    type_map = {"pc": PC, "monitor": Monitor, "docking_station": DockingStation, "phone": Phone,
        "phone_number": PhoneNumber,
        "printer": Printer}
    
    if asset_type not in type_map:
        raise HTTPException(400, "Invalid asset type")
    
    model_class = type_map[asset_type]
    asset = db.query(model_class).filter(model_class.id == asset_id).first()
    if not asset:
        raise HTTPException(404, "Asset not found")
    
    if asset.image_path:
        old_path = os.path.join(UPLOAD_DIR, os.path.basename(asset.image_path))
        if os.path.exists(old_path):
            os.remove(old_path)
        asset.image_path = None
        
        log_audit(db, "image_deleted", asset_type, asset_id,
                  serial_number=asset.serial_number,
                  model=asset.model,
                  details="Image removed")
        db.commit()
    
    return {"message": "Image removed successfully"}


# --- Audit Log Endpoint ---

@router.get("/audit-log")
async def get_audit_log(
    asset_type: Optional[str] = None,
    asset_id: Optional[int] = None,
    action: Optional[str] = None,
    limit: int = Query(default=100, le=500),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db)
):
    """Fetch audit log entries with optional filters."""
    query = db.query(StockAuditLog)
    
    if asset_type:
        query = query.filter(StockAuditLog.asset_type == asset_type)
    if asset_id:
        query = query.filter(StockAuditLog.asset_id == asset_id)
    if action:
        query = query.filter(StockAuditLog.action == action)
    
    total = query.count()
    entries = query.order_by(desc(StockAuditLog.created_at)).offset(offset).limit(limit).all()
    
    return {
        "total": total,
        "entries": [
            {
                "id": e.id,
                "action": e.action,
                "asset_type": e.asset_type,
                "asset_id": e.asset_id,
                "serial_number": e.serial_number,
                "model": e.model,
                "performed_by": e.performed_by,
                "details": e.details,
                "created_at": e.created_at.isoformat() if e.created_at else None
            } for e in entries
        ]
    }


from app.dependencies import get_current_user

@router.post("/discovery/sync")
async def sync_discovery_data(
    payload: DiscoveryPayloadSchema, 
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Synchronize discovered network assets into ITSM."""
    processed = 0
    created = 0
    updated = 0
    processed_monitor_ids = set()

    for device in payload.devices:
        processed += 1
        
        # Determine User
        assigned_user = None
        if device.logged_in_user:
            # Extract username from DOMAIN\username or username@domain
            username = device.logged_in_user
            if "\\" in username:
                username = username.split("\\")[-1]
            elif "@" in username:
                username = username.split("@")[0]
            
            assigned_user = db.query(User).filter(User.username == username).first()

        # 1. Process PC
        pc = db.query(PC).filter(PC.serial_number == device.serial_number).first()
        is_new_pc = pc is None
        
        if is_new_pc:
            pc = PC(
                serial_number=device.serial_number,
                name=device.hostname,
                model=device.model,
                vendor=device.vendor,
                ip_address=device.ip_address,
                mac_address=device.mac_address,
                status="Pending" if not assigned_user else "Assigned",
                created_at=datetime.utcnow(),
                last_seen_at=datetime.utcnow(),
                windows_version=device.windows_version,
                intune_status=device.intune_status,
                antivirus_status=device.antivirus_status,
                ram=device.ram,
                storage=device.storage,
                print_volume_30d=device.print_volume_30d
            )
            db.add(pc)
            db.flush()
            created += 1
            log_audit(db, "created", "pc", pc.id, serial_number=pc.serial_number, model=pc.model, details="Discovered on network")
        else:
            pc.name = device.hostname
            if device.model: pc.model = device.model
            if device.vendor: pc.vendor = device.vendor
            if device.ip_address: pc.ip_address = device.ip_address
            if device.mac_address: pc.mac_address = device.mac_address
            if device.windows_version: pc.windows_version = device.windows_version
            if device.intune_status: pc.intune_status = device.intune_status
            if device.antivirus_status: pc.antivirus_status = device.antivirus_status
            if device.ram: pc.ram = device.ram
            if device.storage: pc.storage = device.storage
            if device.print_volume_30d is not None: pc.print_volume_30d = device.print_volume_30d
            pc.last_seen_at = datetime.utcnow()
            updated += 1
            log_audit(db, "updated", "pc", pc.id, serial_number=pc.serial_number, model=pc.model, details="Updated via network discovery")

        # Process Software
        from app.models.hardware import InstalledSoftware
        
        # 1. Get current active software from DB
        current_software = db.query(InstalledSoftware).filter(
            InstalledSoftware.pc_id == pc.id,
            InstalledSoftware.is_active == True
        ).all()
        
        current_software_dict = {sw.name: sw for sw in current_software}
        incoming_software_dict = {sw.name: sw for sw in device.software}
        
        # 2. Deactivate removed software
        for name, sw_db in current_software_dict.items():
            if name not in incoming_software_dict:
                sw_db.is_active = False
                sw_db.removed_date = datetime.utcnow()
                db.add(sw_db)
                
        # 3. Add new software
        for name, sw_inc in incoming_software_dict.items():
            if name not in current_software_dict:
                new_sw = InstalledSoftware(
                    pc_id=pc.id,
                    name=sw_inc.name,
                    version=sw_inc.version,
                    is_active=True
                )
                db.add(new_sw)
            else:
                # Update version if changed
                if current_software_dict[name].version != sw_inc.version:
                    current_software_dict[name].version = sw_inc.version
                    db.add(current_software_dict[name])
                db.commit()

        # 1c. Process Printers
        from app.models.hardware import InstalledPrinter
        db.query(InstalledPrinter).filter(InstalledPrinter.pc_id == pc.id).delete()
        for p in device.printers:
            new_printer = InstalledPrinter(
                pc_id=pc.id,
                name=p.name,
                driver_name=p.driver_name,
                port_name=p.port_name,
                is_network=p.is_network,
                is_default=p.is_default
            )
            db.add(new_printer)
        db.flush()
        
        db.flush()

        # Handle PC Assignment
        if assigned_user:
            assign_hardware(db, pc, assigned_user, "pc", "Auto-assigned via Discovery Agent")
        elif is_new_pc:
            pc.status = "Pending"
            
        # 2. Process Monitors
        for mon_data in device.monitors:
            query = db.query(Monitor).filter(Monitor.serial_number == mon_data.serial_number)
            if processed_monitor_ids:
                query = query.filter(Monitor.id.notin_(processed_monitor_ids))
            
            monitor = query.first()
            is_new_mon = False
            if not monitor:
                is_new_mon = True
                monitor = Monitor(
                    serial_number=mon_data.serial_number,
                    model=mon_data.model,
                    vendor=mon_data.manufacturer,
                    status="Pending" if not assigned_user else "Assigned",
                    created_at=datetime.utcnow(),
                    last_seen_at=datetime.utcnow(),
                    connected_pc_id=pc.id
                )
                db.add(monitor)
                db.flush()
                created += 1
                log_audit(db, "imported", "monitor", monitor.id, serial_number=monitor.serial_number, model=monitor.model, details=f"Discovered connected to PC {pc.serial_number}")
            else:
                monitor.last_seen_at = datetime.utcnow()
                monitor.connected_pc_id = pc.id
                if mon_data.model: monitor.model = mon_data.model
                if mon_data.manufacturer: monitor.vendor = mon_data.manufacturer
                monitor.last_seen_at = datetime.utcnow()
                updated += 1
                log_audit(db, "updated", "monitor", monitor.id, serial_number=monitor.serial_number, model=monitor.model, details="Updated via network discovery")

            processed_monitor_ids.add(monitor.id)

            # Handle Monitor Assignment
            if assigned_user:
                assign_hardware(db, monitor, assigned_user, "monitor", f"Auto-assigned with PC {device.hostname}")
            elif is_new_mon:
                monitor.status = "Pending"

    for printer in payload.network_printers:
        p_db = db.query(Printer).filter(Printer.serial_number == printer.serial_number).first()
        is_new = p_db is None
        if is_new:
            p_db = Printer(
                serial_number=printer.serial_number,
                model=printer.model,
                ip_address=printer.ip_address,
                mac_address=printer.mac_address,
                status="Available",
                created_at=datetime.utcnow(),
                last_seen_at=datetime.utcnow()
            )
            db.add(p_db)
            db.flush()
            created += 1
            log_audit(db, "created", "printer", p_db.id, serial_number=p_db.serial_number, model=p_db.model, details="Discovered on network")
        else:
            p_db.last_seen_at = datetime.utcnow()
            if printer.model: p_db.model = printer.model
            if printer.ip_address: p_db.ip_address = printer.ip_address
            if printer.mac_address: p_db.mac_address = printer.mac_address
            updated += 1
            log_audit(db, "updated", "printer", p_db.id, serial_number=p_db.serial_number, model=p_db.model, details="Updated via network discovery")

    db.commit()
    return {
        "message": "Discovery sync completed",
        "processed_devices": processed,
        "assets_created": created,
        "assets_updated": updated
    }

@router.put("/stock/{asset_type}/{asset_id}")
async def update_asset_details(asset_type: str, asset_id: int, payload: AssetUpdateSchema, db: Session = Depends(get_db)):
    """Update basic asset details."""
    type_map = {
        "pc": PC, "monitor": Monitor, "docking_station": DockingStation, 
        "phone": Phone, "phone_number": PhoneNumber, "printer": Printer
    }
    if asset_type not in type_map:
        raise HTTPException(400, "Invalid asset type")
        
    model_class = type_map[asset_type]
    asset = db.query(model_class).filter(model_class.id == asset_id).first()
    
    if not asset:
        raise HTTPException(404, "Asset not found")

    changes = []
    
    if payload.name is not None and hasattr(asset, 'name') and asset.name != payload.name:
        changes.append(f"Name: {asset.name} -> {payload.name}")
        asset.name = payload.name
        
    if payload.serial_number is not None and asset.serial_number != payload.serial_number:
        # Check duplicate
        existing = db.query(model_class).filter(model_class.serial_number == payload.serial_number, model_class.id != asset_id).first()
        if existing:
            raise HTTPException(400, "Serial number already in use")
        changes.append(f"S/N: {asset.serial_number} -> {payload.serial_number}")
        asset.serial_number = payload.serial_number
        
    if payload.mac_address is not None and hasattr(asset, 'mac_address') and asset.mac_address != payload.mac_address:
        changes.append(f"MAC: {asset.mac_address} -> {payload.mac_address}")
        asset.mac_address = payload.mac_address
        
    if payload.ip_address is not None and hasattr(asset, 'ip_address') and asset.ip_address != payload.ip_address:
        changes.append(f"IP: {asset.ip_address} -> {payload.ip_address}")
        asset.ip_address = payload.ip_address

    if payload.model is not None and asset.model != payload.model:
        changes.append(f"Model: {asset.model} -> {payload.model}")
        asset.model = payload.model
        
    if payload.phone_number is not None and hasattr(asset, 'phone_number') and asset.phone_number != payload.phone_number:
        changes.append(f"Phone: {asset.phone_number} -> {payload.phone_number}")
        asset.phone_number = payload.phone_number

    if payload.notes is not None and asset.notes != payload.notes:
        changes.append("Notes updated")
        asset.notes = payload.notes
        
    if changes:
        db.add(asset)
        log_audit(db, "updated", asset_type, asset_id, serial_number=asset.serial_number, model=asset.model, details=" | ".join(changes))
        db.commit()
        
    return {"message": "Asset updated successfully"}

