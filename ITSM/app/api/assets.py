"""Asset management API endpoints"""
from fastapi import APIRouter, Depends, HTTPException, Query, status, Request
from app.dependencies import check_permission
from sqlalchemy.orm import Session
from sqlalchemy import or_
from typing import List, Optional
import logging
from datetime import datetime

from app.db import get_db
from app.schemas.asset import (
    AssetCreate, AssetUpdate, AssetResponse,
    AssetType, AssetStatus,
    MaintenanceCreate, MaintenanceResponse
)
from app.models.asset import Asset, AssetMaintenance

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/assets", tags=["assets"], dependencies=[Depends(check_permission("hardware"))])


from app.models.hardware import PC, Monitor, DockingStation, Phone, PhoneNumber
from app.models.user import User


# ============ STATISTICS (must be BEFORE /{asset_id}) ============

@router.get("/stats/summary", tags=["assets"])
async def get_asset_statistics(
    db: Session = Depends(get_db),
):
    """Get asset management statistics"""
    try:
        # Count legacy assets
        legacy_count = db.query(Asset).filter(Asset.is_active == True).count()

        # Count hardware-based assets (users with any hardware)
        users_with_hardware = 0
        unassigned_pcs = 0
        try:
            # Get distinct user IDs that have hardware assigned to eliminate N+1 queries
            assigned_pc_users = set(row[0] for row in db.query(PC.current_user_id).filter(PC.current_user_id.isnot(None)).all())
            assigned_monitor_users = set(row[0] for row in db.query(Monitor.current_user_id).filter(Monitor.current_user_id.isnot(None)).all())
            assigned_dock_users = set(row[0] for row in db.query(DockingStation.current_user_id).filter(DockingStation.current_user_id.isnot(None)).all())
            assigned_phone_users = set(row[0] for row in db.query(Phone.current_user_id).filter(Phone.current_user_id.isnot(None)).all())
            
            all_hardware_users = assigned_pc_users | assigned_monitor_users | assigned_dock_users | assigned_phone_users
            users_with_hardware = len(all_hardware_users)
            
            unassigned_pcs = db.query(PC).filter(PC.current_user_id == None).count()
        except Exception as e:
            logger.error(f"Error counting hardware: {str(e)}")

        total_assets = legacy_count + users_with_hardware + unassigned_pcs

        by_type = {}
        for asset_type in AssetType:
            count = db.query(Asset).filter(
                Asset.asset_type == asset_type,
                Asset.is_active == True
            ).count()
            by_type[asset_type.value] = count
        # Add hardware counts
        by_type["computer"] = by_type.get("computer", 0) + users_with_hardware + unassigned_pcs

        by_status = {}
        for asset_status in AssetStatus:
            count = db.query(Asset).filter(
                Asset.status == asset_status,
                Asset.is_active == True
            ).count()
            by_status[asset_status.value] = count
        # Hardware-based counts
        by_status["in_use"] = by_status.get("in_use", 0) + users_with_hardware
        by_status["available"] = by_status.get("available", 0) + unassigned_pcs

        from app.models.asset import LicenseRegistration
        total_licenses = db.query(LicenseRegistration).filter(
            LicenseRegistration.is_active == True
        ).count()

        return {
            "total_assets": total_assets,
            "by_type": by_type,
            "by_status": by_status,
            "total_licenses": total_licenses,
        }

    except Exception as e:
        logger.error(f"Error getting statistics: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ TAG LOOKUP (must be BEFORE /{asset_id}) ============

@router.get("/tag/{asset_tag}", response_model=AssetResponse)
async def get_asset_by_tag(
    asset_tag: str,
    db: Session = Depends(get_db),
):
    """Get asset by tag"""
    asset = db.query(Asset).filter(
        Asset.asset_tag == asset_tag,
        Asset.is_active == True
    ).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    return asset


# ============ ASSET MANAGEMENT ============

@router.get("/")
async def list_assets(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    asset_type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """List all assets with filtering and search, combining hardware models.
    
    Returns specs in array format: specs.pcs[], specs.monitors[], etc.
    so the frontend can display multiple devices per user.
    """
    logger.info(f"list_assets called with asset_type={asset_type}, status={status}, search={search}")
    try:
        results = []

        try:
            users = db.query(User).all()
        except Exception as e:
            logger.error(f"Error querying users: {str(e)}")
            users = []

        logger.info(f"Found {len(users)} users")

        # Fetch all assigned hardware in 4 queries instead of 4 per user (N+1 fix)
        from collections import defaultdict
        
        assigned_pcs = db.query(PC).filter(PC.current_user_id.isnot(None)).all()
        user_pcs = defaultdict(list)
        for p in assigned_pcs: user_pcs[p.current_user_id].append(p)
            
        assigned_monitors = db.query(Monitor).filter(Monitor.current_user_id.isnot(None)).all()
        user_monitors = defaultdict(list)
        for m in assigned_monitors: user_monitors[m.current_user_id].append(m)
            
        assigned_docks = db.query(DockingStation).filter(DockingStation.current_user_id.isnot(None)).all()
        user_docks = defaultdict(list)
        for d in assigned_docks: user_docks[d.current_user_id].append(d)
            
        assigned_phones = db.query(Phone).filter(Phone.current_user_id.isnot(None)).all()
        user_phones = defaultdict(list)
        for ph in assigned_phones: user_phones[ph.current_user_id].append(ph)

        assigned_phone_numbers = db.query(PhoneNumber).filter(PhoneNumber.current_user_id.isnot(None)).all()
        user_phone_numbers = defaultdict(list)
        for pn in assigned_phone_numbers: user_phone_numbers[pn.current_user_id].append(pn)

        for u in users:
            try:
                pcs = user_pcs[u.id]
                monitors = user_monitors[u.id]
                docks = user_docks[u.id]
                phones = user_phones[u.id]
                phone_numbers = user_phone_numbers[u.id]

                # Build specs with arrays so the frontend can render multiple devices
                specs = {
                    "department": getattr(u, 'department', '') or "",
                    "pcs": [{"id": p_item.id, "name": p_item.name or "", "model": p_item.model or "", "sn": p_item.serial_number or ""} for p_item in pcs],
                    "monitors": [{"id": m_item.id, "model": m_item.model or "", "sn": m_item.serial_number or ""} for m_item in monitors],
                    "docking_sn": docks[0].serial_number if docks else "",
                    "docking_id": docks[0].id if docks else None,
                    "phone_model": phones[0].model if phones else "",
                    "phone_number": phone_numbers[0].phone_number if phone_numbers else "",
                    "phone_id": phones[0].id if phones else None,
                    "phone_number_id": phone_numbers[0].id if phone_numbers else None,
                    "accessories": "",
                    # Keep flat fields for backward compatibility
                    "hostname": pcs[0].name if pcs else "",
                    "laptop_model": pcs[0].model if pcs else "",
                    "laptop_sn": pcs[0].serial_number if pcs else "",
                    "monitor_model": monitors[0].model if monitors else "",
                    "monitor_sn": monitors[0].serial_number if monitors else "",
                }

                import json
                main_pc = pcs[0] if pcs else None

                bundle = {
                    "id": 1000000 + u.id,
                    "name": u.full_name,
                    "description": "",
                    "asset_type": "computer" if pcs else "other",
                    "status": "in_use" if (pcs or monitors or docks or phones) else "available",
                    "asset_tag": str(u.id),
                    "serial_number": main_pc.serial_number if main_pc else "",
                    "model_number": main_pc.model if main_pc else "",
                    "manufacturer": "",
                    "location": "",
                    "assigned_user_id": u.id,
                    "purchase_date": None,
                    "purchase_cost": None,
                    "warranty_expiry": None,
                    "depreciation_rate": None,
                    "specifications": json.dumps(specs),
                    "license_key": None,
                    "license_expiry": None,
                    "end_of_life_date": None,
                    "notes": "",
                    "is_active": True,
                    "created_at": datetime.utcnow().isoformat(),
                    "updated_at": datetime.utcnow().isoformat()
                }

                # Apply filters
                if asset_type and bundle["asset_type"] != asset_type:
                    continue
                if status and bundle["status"] != status:
                    continue

                # Apply search filter
                if search:
                    search_lower = search.lower()
                    searchable = f"{u.full_name} {specs.get('hostname', '')} {specs.get('laptop_sn', '')} {specs.get('monitor_sn', '')} {specs.get('docking_sn', '')} {specs.get('phone_number', '')} {specs.get('phone_model', '')} {specs.get('department', '')}".lower()
                    if search_lower not in searchable:
                        continue

                results.append(bundle)
            except Exception as e:
                logger.error(f"Error processing user {u.id}: {str(e)}")
                continue

        # Include unassigned PCs
        unassigned_pcs = db.query(PC).filter(PC.current_user_id == None).all()
        for un_pc in unassigned_pcs:
            try:
                import json
                specs = {
                    "pcs": [{"id": un_pc.id, "name": un_pc.name or "", "model": un_pc.model or "", "sn": un_pc.serial_number or ""}],
                    "monitors": [],
                    "docking_sn": "",
                    "phone_model": "",
                    "phone_number": "",
                    "accessories": "",
                    "hostname": un_pc.name or "",
                    "laptop_model": un_pc.model or "",
                    "laptop_sn": un_pc.serial_number or "",
                }
                bundle = {
                    "id": 1000000 + un_pc.id, 
                    "name": "Unassigned PC", 
                    "description": "", 
                    "asset_type": "computer",
                    "status": "available", 
                    "asset_tag": f"PC-{un_pc.id}", 
                    "serial_number": un_pc.serial_number,
                    "model_number": un_pc.model, 
                    "manufacturer": "", 
                    "location": "", 
                    "assigned_user_id": None,
                    "purchase_date": None, 
                    "purchase_cost": None, 
                    "warranty_expiry": None, 
                    "depreciation_rate": None,
                    "specifications": json.dumps(specs), 
                    "license_key": None, 
                    "license_expiry": None,
                    "end_of_life_date": None, 
                    "notes": "", 
                    "is_active": True, 
                    "created_at": datetime.utcnow().isoformat(),
                    "updated_at": datetime.utcnow().isoformat()
                }

                if asset_type and bundle["asset_type"] != asset_type:
                    continue
                if status and bundle["status"] != status:
                    continue

                if search:
                    search_lower = search.lower()
                    searchable = f"Unassigned PC {un_pc.name or ''} {un_pc.serial_number or ''} {un_pc.model or ''}".lower()
                    if search_lower not in searchable:
                        continue

                results.append(bundle)
            except Exception as e:
                logger.error(f"Error processing unassigned PC {un_pc.id}: {str(e)}")
                continue

        # Include legacy manual assets
        try:
            legacy_assets = db.query(Asset).filter(Asset.is_active == True).all()
            for la in legacy_assets:
                la_status = la.status.value if hasattr(la.status, "value") else la.status
                la_type = la.asset_type.value if hasattr(la.asset_type, "value") else la.asset_type
                
                if asset_type and la_type != asset_type:
                    continue
                if status and la_status != status:
                    continue

                if search:
                    search_lower = search.lower()
                    searchable = f"{la.name or ''} {la.serial_number or ''} {la.asset_tag or ''} {la.model_number or ''}".lower()
                    if search_lower not in searchable:
                        continue

                bundle = {
                    "id": la.id, 
                    "name": la.name, 
                    "description": la.description, 
                    "asset_type": la.asset_type,
                    "status": la.status, 
                    "asset_tag": la.asset_tag, 
                    "serial_number": la.serial_number,
                    "model_number": la.model_number, 
                    "manufacturer": la.manufacturer, 
                    "location": la.location, 
                    "assigned_user_id": la.assigned_user_id,
                    "purchase_date": la.purchase_date.isoformat() if la.purchase_date else None, 
                    "purchase_cost": la.purchase_cost, 
                    "warranty_expiry": la.warranty_expiry.isoformat() if la.warranty_expiry else None, 
                    "depreciation_rate": la.depreciation_rate,
                    "specifications": la.specifications, 
                    "license_key": la.license_key, 
                    "license_expiry": la.license_expiry.isoformat() if la.license_expiry else None,
                    "end_of_life_date": la.end_of_life_date.isoformat() if la.end_of_life_date else None, 
                    "notes": la.notes, 
                    "is_active": la.is_active, 
                    "created_at": la.created_at.isoformat() if la.created_at else datetime.utcnow().isoformat(),
                    "updated_at": la.updated_at.isoformat() if la.updated_at else datetime.utcnow().isoformat()
                }
                results.append(bundle)
        except Exception as e:
            logger.error(f"Error processing legacy assets: {str(e)}")

        return results
    except Exception as e:
        logger.error(f"Error in list_assets: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error loading assets: {str(e)}")


@router.get("/export")
async def export_assets(
    asset_type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Export all assets to Excel"""
    import io
    import json
    import pandas as pd
    from fastapi.responses import StreamingResponse
    
    try:
        # Re-use list_assets logic to get the filtered list
        results = await list_assets(
            skip=0, limit=10000, 
            asset_type=asset_type, 
            status=status, 
            search=search, 
            db=db
        )
        
        export_data = []
        for r in results:
            specs_str = r.get("specifications", "{}")
            try:
                specs = json.loads(specs_str) if isinstance(specs_str, str) else specs_str
                if not isinstance(specs, dict):
                    specs = {}
            except:
                specs = {}
                
            pcs = specs.get("pcs", [])
            monitors = specs.get("monitors", [])
            
            num_rows = max(len(pcs), len(monitors), 1)
            
            for i in range(num_rows):
                # Fill down PC: if we run out of PCs but have monitors, assign them to the last available PC
                pc = pcs[i] if i < len(pcs) else (pcs[-1] if pcs else {})
                mon = monitors[i] if i < len(monitors) else {}
                
                export_data.append({
                    "ID": r.get("id", ""),
                    "Name / Assigned User": r.get("name", ""),
                    "Status": r.get("status", ""),
                    "Asset Type": r.get("asset_type", ""),
                    "PC Hostname": pc.get("name", ""),
                    "PC Model": pc.get("model", ""),
                    "PC Serial Number": pc.get("sn", ""),
                    "Monitor Model": mon.get("model", ""),
                    "Monitor Serial Number": mon.get("sn", ""),
                    "Docking Station": specs.get("docking_sn", "") if i == 0 else "",
                    "Phone Model": specs.get("phone_model", "") if i == 0 else "",
                    "Phone Number": specs.get("phone_number", "") if i == 0 else "",
                    "Accessories": specs.get("accessories", "") if i == 0 else "",
                    "Department": specs.get("department", "")
                })
            
        df_dashboard = pd.DataFrame(export_data)
        df = df_dashboard.copy()
        if 'Status' in df:
            df = df.drop(columns=['Status'])
            
        output = io.BytesIO()
        
        import openpyxl
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.chart import PieChart, BarChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 1. Write the raw data
            df.to_excel(writer, index=False, sheet_name='Data', startrow=1)
            workbook = writer.book
            data_sheet = writer.sheets['Data']
            
            # Format Data sheet
            data_sheet.title = "Asset Database"
            data_sheet.sheet_view.showGridLines = False
            
            # Add Title to Data sheet
            data_sheet['A1'] = "Detailed Hardware Inventory Database"
            data_sheet['A1'].font = Font(size=16, bold=True, color="1F4E78")
            
            # Auto width & Freeze panes
            for col_idx, column_cells in enumerate(data_sheet.columns, 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 4, 45)  # Add more padding
                data_sheet.column_dimensions[col_letter].width = adjusted_width
                
            data_sheet.freeze_panes = "A3" # Freeze header row
            
            # Add table formatting
            if len(df) > 0:
                tab = Table(displayName="AssetsDatabase", ref=f"A2:{data_sheet.cell(row=len(df)+2, column=len(df.columns)).coordinate}")
                # Use a very clean, light styling
                style = TableStyleInfo(name="TableStyleLight8", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tab.tableStyleInfo = style
                data_sheet.add_table(tab)
                
                # Professional Row Heights and Alignment
                data_sheet.row_dimensions[2].height = 25  # Header row height
                for row in range(3, len(df)+3):
                    data_sheet.row_dimensions[row].height = 22
                    # Center align ID and Asset Type (which is now column 3)
                    data_sheet.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                    data_sheet.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='center')
                    # Left align the rest but center vertically
                    for col in range(2, len(df.columns) + 1):
                        if col not in [1, 3]:
                            data_sheet.cell(row=row, column=col).alignment = Alignment(vertical='center')

            # 2. Create Executive Summary Dashboard Sheet
            summary_sheet = workbook.create_sheet('Executive Summary') # Append to the end
            workbook.active = data_sheet # Make Asset Database the default active sheet
            summary_sheet.sheet_view.showGridLines = False
            
            # Modern Header
            summary_sheet['B2'] = " KOSTAL ITSM"
            summary_sheet['B2'].font = Font(size=24, bold=True, color="FFFFFF")
            summary_sheet['B2'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            summary_sheet['B2'].alignment = Alignment(horizontal="left", vertical="center")
            
            summary_sheet['E2'] = "HARDWARE INVENTORY REPORT "
            summary_sheet['E2'].font = Font(size=14, bold=True, color="FFFFFF")
            summary_sheet['E2'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            summary_sheet['E2'].alignment = Alignment(horizontal="right", vertical="center")
            
            summary_sheet.merge_cells('B2:D3')
            summary_sheet.merge_cells('E2:K3')
            
            # Date Generated
            from datetime import datetime
            summary_sheet['B4'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            summary_sheet['B4'].font = Font(italic=True, color="7F7F7F")
            
            # KPI Cards
            kpi_titles = ["Total Assets", "Active in Use", "Available / Spares", "Maintenance Required"]
            
            total_assets = len(df_dashboard)
            in_use = len(df_dashboard[df_dashboard['Status'].str.lower() == 'in_use']) if 'Status' in df_dashboard else 0
            available = len(df_dashboard[df_dashboard['Status'].str.lower() == 'available']) if 'Status' in df_dashboard else 0
            maintenance = len(df_dashboard[df_dashboard['Status'].str.lower().isin(['damaged', 'maintenance'])]) if 'Status' in df_dashboard else 0
            
            kpi_values = [total_assets, in_use, available, maintenance]
            cols = ['B', 'E', 'H', 'J']
            
            thin_border = Border(left=Side(style='thin', color="CCCCCC"), 
                                 right=Side(style='thin', color="CCCCCC"), 
                                 top=Side(style='thin', color="CCCCCC"), 
                                 bottom=Side(style='thin', color="CCCCCC"))
            
            for i in range(4):
                col = cols[i]
                end_col = chr(ord(col) + 1) if i < 3 else 'K'
                
                # Title
                summary_sheet[f'{col}6'] = kpi_titles[i]
                summary_sheet[f'{col}6'].font = Font(size=11, color="7F7F7F", bold=True)
                summary_sheet[f'{col}6'].alignment = Alignment(horizontal="center", vertical="center")
                summary_sheet[f'{col}6'].fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
                
                # Value
                summary_sheet[f'{col}7'] = kpi_values[i]
                summary_sheet[f'{col}7'].font = Font(size=24, bold=True, color="2E75B6")
                summary_sheet[f'{col}7'].alignment = Alignment(horizontal="center", vertical="center")
                summary_sheet[f'{col}7'].fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
                
                summary_sheet.merge_cells(f'{col}6:{end_col}6')
                summary_sheet.merge_cells(f'{col}7:{end_col}8')
                
                # Apply borders
                for r_idx in range(6, 9):
                    for c_idx in range(openpyxl.utils.column_index_from_string(col), openpyxl.utils.column_index_from_string(end_col)+1):
                        summary_sheet.cell(row=r_idx, column=c_idx).border = thin_border

            # Put data used for charts far out of sight (Column CV / 100) so charts render correctly
            data_start_col = 100 
            status_counts = df_dashboard['Status'].value_counts() if 'Status' in df_dashboard else pd.Series()
            
            summary_sheet.cell(row=1, column=data_start_col).value = "Status"
            summary_sheet.cell(row=1, column=data_start_col+1).value = "Count"
            row_idx = 2
            for status, count in status_counts.items():
                summary_sheet.cell(row=row_idx, column=data_start_col).value = status if status else "Unknown"
                summary_sheet.cell(row=row_idx, column=data_start_col+1).value = count
                row_idx += 1
                
            # Create Pie Chart
            if not status_counts.empty and row_idx > 2:
                pie = PieChart()
                pie.title = "Equipment Status Distribution"
                pie.style = 2  
                labels = Reference(summary_sheet, min_col=data_start_col, min_row=2, max_row=row_idx-1)
                data = Reference(summary_sheet, min_col=data_start_col+1, min_row=1, max_row=row_idx-1)
                pie.add_data(data, titles_from_data=True)
                pie.set_categories(labels)
                pie.dataLabels = DataLabelList()
                pie.dataLabels.showPercent = True
                pie.width = 14
                pie.height = 9
                summary_sheet.add_chart(pie, "B11")
                
            # Equipment Assignment Stats
            if 'PC Hostname' in df:
                has_pc = df['PC Hostname'].apply(lambda x: "Assigned" if str(x).strip() else "Not Assigned").value_counts()
                summary_sheet.cell(row=1, column=data_start_col+3).value = "Assignment"
                summary_sheet.cell(row=1, column=data_start_col+4).value = "Count"
                
                r2 = 2
                for cat, count in has_pc.items():
                    summary_sheet.cell(row=r2, column=data_start_col+3).value = cat
                    summary_sheet.cell(row=r2, column=data_start_col+4).value = count
                    r2 += 1
                    
                if not has_pc.empty and r2 > 2:
                    bar = BarChart()
                    bar.title = "PC Allocation"
                    bar.style = 13
                    labels2 = Reference(summary_sheet, min_col=data_start_col+3, min_row=2, max_row=r2-1)
                    data2 = Reference(summary_sheet, min_col=data_start_col+4, min_row=1, max_row=r2-1)
                    bar.add_data(data2, titles_from_data=True)
                    bar.set_categories(labels2)
                    bar.varyColors = True
                    bar.legend = None
                    bar.width = 14
                    bar.height = 9
                    summary_sheet.add_chart(bar, "G11")
                
            # Clean up column widths
            summary_sheet.column_dimensions['A'].width = 3
            summary_sheet.column_dimensions['B'].width = 15
            summary_sheet.column_dimensions['C'].width = 10
            summary_sheet.column_dimensions['D'].width = 15
            summary_sheet.column_dimensions['E'].width = 15
            summary_sheet.column_dimensions['F'].width = 10
            summary_sheet.column_dimensions['G'].width = 15
            summary_sheet.column_dimensions['H'].width = 15
            summary_sheet.column_dimensions['I'].width = 10
            summary_sheet.column_dimensions['J'].width = 15
            summary_sheet.column_dimensions['K'].width = 15
            
            
        output.seek(0)
        
        headers = {
            'Content-Disposition': 'attachment; filename="hardware_inventory.xlsx"'
        }
        return StreamingResponse(
            output, 
            headers=headers, 
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Error exporting assets: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error exporting assets: {str(e)}")


@router.get("/{asset_id}", response_model=AssetResponse)
async def get_asset(
    asset_id: int,
    db: Session = Depends(get_db),
):
    """Get specific asset details"""
    if asset_id >= 1000000:
        raise HTTPException(status_code=404, detail="Hardware bundle pseudo-record is not found in Asset table")

    asset = db.query(Asset).filter(Asset.id == asset_id, Asset.is_active == True).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    return asset


@router.post("/", response_model=AssetResponse, status_code=201)
async def create_asset(
    asset_create: AssetCreate,
    db: Session = Depends(get_db),
):
    """Create a new asset"""
    try:
        # Check if asset tag already exists and is active
        existing = db.query(Asset).filter(Asset.asset_tag == asset_create.asset_tag, Asset.is_active == True).first()
        if existing:
            raise HTTPException(
                status_code=400,
                detail=f"Asset tag {asset_create.asset_tag} already exists"
            )

        asset = Asset(
            **asset_create.model_dump()
        )
        db.add(asset)
        db.commit()
        db.refresh(asset)

        logger.info(f"Asset created: {asset.name} ({asset.asset_tag})")
        return asset

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error creating asset: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/{asset_id}", response_model=AssetResponse)
async def update_asset(
    asset_id: int,
    asset_update: AssetUpdate,
    db: Session = Depends(get_db),
):
    """Update an asset"""
    if asset_id >= 1000000:
        raise HTTPException(status_code=400, detail="Cannot edit Excel-imported hardware from this form.")

    try:
        asset = db.query(Asset).filter(Asset.id == asset_id).first()
        if not asset:
            raise HTTPException(status_code=404, detail="Asset not found")

        # Update only provided fields
        update_data = asset_update.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(asset, field, value)

        db.commit()
        db.refresh(asset)

        logger.info(f"Asset updated: {asset.name}")
        return asset

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error updating asset: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/{asset_id}", status_code=204)
async def delete_asset(
    asset_id: int,
    db: Session = Depends(get_db),
):
    """Soft delete an asset"""
    if asset_id >= 1000000:
        raise HTTPException(status_code=400, detail="Hardware bundles imported from Excel cannot be deleted here yet.")

    try:
        asset = db.query(Asset).filter(Asset.id == asset_id).first()
        if not asset:
            raise HTTPException(status_code=404, detail="Asset not found")

        asset.is_active = False
        asset.asset_tag = f"{asset.asset_tag}_deleted_{asset.id}"
        db.commit()

        logger.info(f"Asset deleted: {asset.name}")

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error deleting asset: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ MAINTENANCE MANAGEMENT ============

@router.post("/{asset_id}/maintenance", response_model=MaintenanceResponse, status_code=201)
async def record_maintenance(
    asset_id: int,
    maintenance: MaintenanceCreate,
    db: Session = Depends(get_db),
):
    """Record maintenance for an asset"""
    try:
        # Verify asset exists
        asset = db.query(Asset).filter(Asset.id == asset_id).first()
        if not asset:
            raise HTTPException(status_code=404, detail="Asset not found")

        maintenance_record = AssetMaintenance(
            asset_id=asset_id,
            **maintenance.model_dump(exclude={'asset_id'})
        )
        db.add(maintenance_record)
        db.commit()
        db.refresh(maintenance_record)

        logger.info(f"Maintenance recorded for asset {asset_id}")
        return maintenance_record

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error recording maintenance: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{asset_id}/maintenance", response_model=List[MaintenanceResponse])
async def get_asset_maintenance(
    asset_id: int,
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    db: Session = Depends(get_db),
):
    """Get maintenance history for an asset"""
    # Verify asset exists
    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    records = db.query(AssetMaintenance).filter(
        AssetMaintenance.asset_id == asset_id
    ).order_by(AssetMaintenance.maintenance_date.desc()).offset(skip).limit(limit).all()

    return records
